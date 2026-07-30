"""Template-format workbook exporter.

Writes the graph as one sheet per project plus a Planner sheet, matching the
template contract: header-name columns, explicit edges in 'Depends on'
(task-level sources prefixed 'task: '), Seq written only where order is
user-asserted (parallel ranks), refs stamped, and dependency-smelling notes
marked in 'Proposed: Depends on' with an INVESTIGATE sentinel for later
human/LLM adjudication. Start/Priority/Follow-up/Remind carry the planner
fields and Today carries current list membership (1-based order), so an
export -> re-import round trip is lossless. Deterministic: identical graph
and Today list -> identical cells.

`build_export` is pure; `write_workbook` is the thin openpyxl shell.
"""
from __future__ import annotations

from collections import defaultdict

from . import recurrence
from .graph import Graph
from .importer import generate_ref, note_dependency_terms
from .model import Node

PROJECT_HEADER = (
    "Project", "Milestone", "Goal", "Task", "Seq", "Depends on",
    "Proposed: Depends on", "Start", "Wait reason", "Deadline", "Est (min)",
    "Priority", "Follow-up (days)", "Remind", "Repeat", "Today", "Steps",
    "Links", "Tags", "Notes", "Ref",
)
PLANNER_HEADER = (
    "Task", "Start", "Wait reason", "Deadline", "Est (min)", "Priority",
    "Follow-up (days)", "Remind", "Repeat", "Today", "Steps", "Links",
    "Tags", "Notes", "Ref",
)

INVESTIGATE_PREFIX = "INVESTIGATE"

# The colour legend, written back on the name cell so a round trip keeps the
# convention the user actually reads. 'not_begun' IS painted (yellow): it means
# the user explicitly marked the row "not begun but will land this quarter",
# which is a statement. A node with `health is None` was never coloured and is
# left unpainted, so an untouched sheet still round-trips looking untouched.
LEGEND_FILL = {
    "on_track": "70AD47",
    "not_begun": "FFC000",
    "off_track": "5B9BD5",
    "wont_finish": "FF0000",
}


# --- the project manager's format (the deliverable) -----------------------
#
# Measured from the user's own template. This is a *deliverable*, not an
# interchange format: it has no Ref, Seq, Depends on, Est, Tags, Steps, Links
# or Today column, so an export in this shape is deliberately LOSSY and cannot
# round-trip everything. `build_export` above stays the lossless format.

PM_HEADER = (
    "Project Milestone(s)", "Goal(s)", "Tasks", "Team Lead",
    "Responsible Party", "Success Criteria", "Start", "Finish",
    "Priority", "Trouble shooting Comments", "Notes",
)

# (row, text) in column F, above the header. Row 5 is left blank and the last
# line sits on row 6, matching the live tracker rather than the template (the
# template has no blank row and puts its header on row 6). Import finds the
# header by NAME, so either layout reads back.
PM_LEGEND = (
    (1, "Green - on track for deadline"),
    (2, "Yellow – tasks that have not begun but will be completed by "
        "the quarter"),
    (3, "Red – tasks that will not be done this quarter"),
    (4, "Blue – off track & Blue w/ Strikethrough – off track but was "
        "completed"),
    (6, "Green & Strikethrough - completed (by deadline)"),
)
PM_HEADER_ROW = 7
PM_PREAMBLE_LABELS = (
    "Project name", "Tier Level", "Project start date", "Project finish date",
)
# the structural column each kind is written into (0-based); a project has no
# row of its own — it is named in the preamble
PM_COL_BY_KIND = {"milestone": 0, "goal": 1, "task": 2}
# what the workbook cannot carry, reported so the loss is never silent
PM_OMITTED_FIELDS = (
    "dependencies", "sequence ranks", "estimates", "tags", "steps", "links",
    "Today membership", "recurrence rules", "wait reasons", "completion dates",
)


def _sheet_title(name: str) -> str:
    cleaned = "".join(ch for ch in name if ch not in "[]:*?/\\")
    return cleaned[:31] or "Sheet"


def steps_cell_text(steps: list[dict]) -> str | None:
    """Steps -> '[x] cut; [ ] deburr'. The importer splits on the checkbox
    markers themselves, so step names may contain semicolons safely."""
    if not steps:
        return None
    return "; ".join(
        f"[{'x' if s['done'] else ' '}] {s['name']}" for s in steps
    )


def links_cell_text(links: list) -> str | None:
    """Links -> 'label|href; href2' (label omitted when it IS the href)."""
    if not links:
        return None
    return "; ".join(
        l["href"] if l.get("label") in (None, l["href"])
        else f"{l['label']}|{l['href']}"
        for l in links
    )


def build_export(
    g: Graph,
    today_pos: dict[int, int] | None = None,
    steps: dict[int, list[dict]] | None = None,
) -> tuple[list[tuple], int]:
    """Returns ([(sheet_title, header, rows, styles), ...], node_count).

    `styles` runs parallel to `rows`: each entry is None, or a dict naming the
    cell that carries the legend ({'col', 'health', 'done'}). Keeping it beside
    the values rather than inside them leaves this function pure data.

    `today_pos` maps task id -> 1-based position on the Today list; those
    positions are written to the Today column so list membership and order
    survive a round trip. `steps` maps task id -> checklist rows. Both live
    outside the graph, so the caller resolves them and this function stays
    pure."""
    today_pos = today_pos or {}
    steps = steps or {}
    incoming = defaultdict(list)  # to_id -> [Dependency]
    for d in g.deps:
        incoming[d.to_id].append(d)

    def _labels(dep_list):
        out = []
        for d in sorted(dep_list, key=lambda d: d.from_id):
            s = g.nodes[d.from_id]
            out.append(f"task: {s.name}" if s.kind == "task" else s.name)
        return out

    refs: dict[int, str] = {}

    def ref_of(n: Node, parent_ref: str | None) -> str:
        if n.id in refs:
            return refs[n.id]
        ref = n.ref
        if ref is None:
            same_kind = (
                [s for s in g.children(n.parent_id) if s.kind == n.kind]
                if n.parent_id is not None
                else [n]
            )
            ordinal = [s.id for s in same_kind].index(n.id) + 1
            ref = generate_ref(n, parent_ref, ordinal)
        refs[n.id] = ref
        return ref

    def depends_cell(n: Node) -> str | None:
        accepted = [
            d for d in incoming.get(n.id, [])
            if not (d.note or "").startswith("llm:")
        ]
        return "; ".join(_labels(accepted)) or None

    def proposed_cell(n: Node) -> str | None:
        # LLM-proposed deps go here for review (kept = accepted on re-import)
        proposed = [
            d for d in incoming.get(n.id, [])
            if (d.note or "").startswith("llm:")
        ]
        if proposed:
            return "; ".join(_labels(proposed))
        if incoming.get(n.id):
            return None  # already has explicit edges; nothing to investigate
        terms = note_dependency_terms(n.description)
        if not terms:
            return None
        return f"{INVESTIGATE_PREFIX} ({', '.join(terms)}): see Notes"

    def tags_cell(n: Node) -> str | None:
        return ";".join(n.tags) if n.tags else None

    def repeat_cell(n: Node) -> str | None:
        rule = recurrence.loads(n.repeat)
        return recurrence.format_rule(rule) if rule else None

    node_count = 0
    sheets: list[tuple] = []
    projects = sorted(
        (n for n in g.nodes.values() if n.kind == "project"), key=lambda n: n.id
    )
    for project in projects:
        rows: list[tuple] = []
        styles: list[dict | None] = []
        p_ref = ref_of(project, None)

        def row(n: Node, col: int, parent_ref: str | None, seq=None):
            cells = [None] * len(PROJECT_HEADER)
            is_task = n.kind == "task"
            cells[col] = n.name
            cells[4] = seq
            cells[5] = depends_cell(n)
            cells[6] = proposed_cell(n)
            cells[7] = n.earliest_start
            # tasks only, matching the model: containers never carry these
            cells[8] = n.wait_reason if is_task else None
            cells[9] = n.deadline
            cells[10] = n.est_minutes
            cells[11] = n.priority if is_task else None
            cells[12] = n.followup_days if is_task else None
            cells[13] = (1 if n.remind else None) if is_task else None
            cells[14] = repeat_cell(n) if is_task else None
            cells[15] = today_pos.get(n.id) if is_task else None
            cells[16] = steps_cell_text(steps.get(n.id, [])) if is_task else None
            cells[17] = links_cell_text(n.links) if is_task else None
            cells[18] = tags_cell(n)
            cells[19] = n.description
            cells[20] = ref_of(n, parent_ref)
            rows.append(tuple(cells))
            styles.append(
                {"col": col, "health": n.health, "done": n.status == "done"}
                if n.kind == "task"
                else None
            )

        node_count += 1
        row(project, 0, None)
        for milestone in g.children(project.id):
            node_count += 1
            row(milestone, 1, p_ref)
            m_ref = refs[milestone.id]
            for goal in g.children(milestone.id):
                node_count += 1
                row(goal, 2, m_ref)
                g_ref = refs[goal.id]
                for task in g.tasks_under(goal.id):
                    node_count += 1
                    seq = task.seq_index if task.seq_source == "user" else None
                    row(task, 3, g_ref, seq=seq)
        sheets.append((_sheet_title(project.name), PROJECT_HEADER, rows, styles))

    planner = sorted(
        (
            n for n in g.nodes.values()
            if n.kind == "task" and n.parent_id is None
        ),
        key=lambda n: n.id,
    )
    if planner:
        rows, styles = [], []
        for task in planner:
            node_count += 1
            rows.append((
                task.name, task.earliest_start, task.wait_reason,
                task.deadline, task.est_minutes, task.priority,
                task.followup_days, 1 if task.remind else None,
                repeat_cell(task), today_pos.get(task.id),
                steps_cell_text(steps.get(task.id, [])),
                links_cell_text(task.links),
                tags_cell(task), task.description, ref_of(task, None),
            ))
            styles.append({
                "col": 0, "health": task.health, "done": task.status == "done",
            })
        sheets.append(("Planner", PLANNER_HEADER, rows, styles))
    return sheets, node_count


def build_pm_export(g: Graph) -> tuple[list[tuple], int, list[str]]:
    """Returns ([(title, preamble, rows, styles), ...], node_count, omissions).

    One sheet per project in the PM's own layout. `preamble` is the four
    labelled cells above the header; `rows` are 11-tuples matching PM_HEADER;
    `styles` runs parallel to rows, naming the cell that carries the legend.

    Pure — no openpyxl, no I/O — so the layout is unit-testable without a file.
    """
    def date_or_na(value: str | None) -> str:
        return value or "N/A"

    def row_for(n: Node, col: int) -> tuple:
        cells = [None] * len(PM_HEADER)
        cells[col] = n.name
        cells[3] = n.team_lead
        cells[4] = n.responsible_party
        cells[5] = n.success_criteria
        cells[6] = n.earliest_start
        cells[7] = n.deadline
        cells[8] = n.priority if n.kind == "task" else None
        cells[9] = n.troubleshooting
        cells[10] = n.description
        return tuple(cells)

    def style_for(n: Node, col: int) -> dict:
        """Strikethrough only for what is genuinely finished, so a re-import
        reproduces the same states. A container struck here is one the user
        explicitly finished — derived completion is left unstruck, because
        promoting it would turn "all its tasks happen to be done" into the
        stronger claim "I closed this question"."""
        if n.kind == "task":
            done = n.status == "done"
            health = n.health
        else:
            done = n.status == "done"
            # a pivot is not a completion: show it as "will not be done"
            health = "wont_finish" if n.status == "abandoned" else n.health
        return {"col": col, "health": health, "done": done}

    node_count = 0
    sheets: list[tuple] = []
    projects = sorted(
        (n for n in g.nodes.values() if n.kind == "project"), key=lambda n: n.id
    )
    for project in projects:
        rows: list[tuple] = []
        styles: list[dict] = []
        node_count += 1
        for milestone in g.children(project.id):
            node_count += 1
            rows.append(row_for(milestone, PM_COL_BY_KIND["milestone"]))
            styles.append(style_for(milestone, PM_COL_BY_KIND["milestone"]))
            for goal in g.children(milestone.id):
                node_count += 1
                rows.append(row_for(goal, PM_COL_BY_KIND["goal"]))
                styles.append(style_for(goal, PM_COL_BY_KIND["goal"]))
                for task in g.tasks_under(goal.id):
                    node_count += 1
                    rows.append(row_for(task, PM_COL_BY_KIND["task"]))
                    styles.append(style_for(task, PM_COL_BY_KIND["task"]))
        preamble = (
            project.name,
            project.tier_level,
            date_or_na(project.earliest_start),
            date_or_na(project.deadline),
        )
        sheets.append((_sheet_title(project.name), preamble, rows, styles))

    omissions = []
    planner = [
        n for n in g.nodes.values()
        if n.kind == "task" and n.parent_id is None
    ]
    if planner:
        omissions.append(
            f"{len(planner)} standalone planner task(s) have no place in this "
            "format and were not written"
        )
    omissions.append(
        "this format carries no " + ", ".join(PM_OMITTED_FIELDS)
        + " — use the default format for a lossless backup"
    )
    return sheets, node_count, omissions


def write_pm_workbook(path: str, sheets: list[tuple]) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if not sheets:
        # No projects yet — a workbook with zero sheets cannot be saved, and a
        # blank form is the useful thing to hand back anyway.
        sheets = [("Tracker", (None, None, None, None), [], [])]
    for title, preamble, rows, styles in sheets:
        ws = wb.create_sheet(title=title)
        for i, label in enumerate(PM_PREAMBLE_LABELS, start=1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            value = preamble[i - 1]
            if value is not None:
                ws.cell(row=i, column=2, value=value)
        for r, text in PM_LEGEND:
            ws.cell(row=r, column=6, value=text)
        for c, head in enumerate(PM_HEADER, start=1):
            ws.cell(row=PM_HEADER_ROW, column=c, value=head).font = Font(bold=True)
        for offset, (row, style) in enumerate(zip(rows, styles), start=1):
            r = PM_HEADER_ROW + offset
            for c, value in enumerate(row, start=1):
                if value is not None:
                    ws.cell(row=r, column=c, value=value)
            cell = ws.cell(row=r, column=style["col"] + 1)
            hexval = LEGEND_FILL.get(style["health"])
            if hexval:
                cell.fill = PatternFill(
                    "solid", start_color=hexval, end_color=hexval
                )
            if style["done"]:
                cell.font = Font(strikethrough=True)
        ws.freeze_panes = f"A{PM_HEADER_ROW + 1}"
    wb.save(path)


def write_workbook(path: str, sheets: list[tuple]) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if not sheets:  # an empty graph must still produce a loadable file
        sheets = [("Tracker", PROJECT_HEADER, [], [])]
    for title, header, rows, styles in sheets:
        ws = wb.create_sheet(title=title)
        ws.append(list(header))
        for row, style in zip(rows, styles):
            ws.append(list(row))
            if not style:
                continue
            cell = ws.cell(row=ws.max_row, column=style["col"] + 1)
            hexval = LEGEND_FILL.get(style["health"])
            if hexval:
                cell.fill = PatternFill(
                    "solid", start_color=hexval, end_color=hexval
                )
            if style["done"]:
                cell.font = Font(strikethrough=True)
    wb.save(path)
