"""End-to-end test of the React dashboard against the real command layer.

Playwright cannot drive a Tauri window, but it does not need to: the frontend
is deliberately logic-free and its only door to the application is
`window.__TAURI__.core.invoke` (commands "pt", "env_info", "graph_html").
Here that door is wired to a Python function that runs the *real* CLI against
a scratch database -- so this exercises the same argv forms and the same JSON
the desktop app consumes, and a regression in either the UI or the command
layer fails the test.

The built app (app/dist) is an ES-module bundle, which chromium refuses to
load from file:// (module scripts require CORS), so a throwaway local HTTP
server serves the dist directory instead.

Run:  python -m pytest tests/e2e -q          (skips if playwright is absent)
"""
from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import threading
from datetime import date, timedelta
from functools import partial
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

import pytest

pytest.importorskip("playwright", reason="playwright is not installed")
from playwright.sync_api import expect, sync_playwright  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
DIST = ROOT / "app" / "dist"
INDEX = DIST / "index.html"

if not INDEX.exists():
    pytest.skip(
        "app/dist/index.html is missing — run: npm --prefix app/ui run build",
        allow_module_level=True,
    )

expect.set_options(timeout=10_000)

GREEN = "rgb(12, 163, 12)"  # --good: #0ca30c

# A synthetic board with every health state present, so the legend, the
# stacked bar and the impact ranking all have something to render.
FIXTURE = [
    ("add", "project", "Bridge Retrofit"),
    ("add", "milestone", "Construction", "--parent", "1"),
    ("add", "goal", "Steelwork", "--parent", "2"),
    ("add", "task", "Survey the span", "--parent", "3"),
    ("add", "task", "Order steel", "--parent", "3"),
    ("add", "task", "Erect falsework", "--parent", "3"),
    ("add", "project", "Harbour Lights"),
    ("add", "milestone", "Design", "--parent", "7"),
    ("add", "goal", "Optics", "--parent", "8"),
    ("add", "task", "Lens spec", "--parent", "9"),
    ("add", "project", "Greenfield"),  # deliberately empty
    # a planner task that carries a dependency: the DAG view must render it
    # as a dashed .dcard.planner card (task 6 -> task 12, a task-level edge)
    ("add", "task", "Calibrate the strain rig"),
    ("dep", "add", "6", "12"),
]
HEALTHS = {4: "on_track", 5: "not_begun", 6: "off_track", 10: "wont_finish"}


def run_cli(db: Path, args: list[str]) -> object:
    proc = subprocess.run(
        [sys.executable, "-m", "protracker.cli", "--db", str(db), "--json", *args],
        cwd=ROOT, capture_output=True, text=True,
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stdout.strip() or proc.stderr.strip())
    return json.loads(proc.stdout) if proc.stdout.strip() else None


@pytest.fixture(scope="module")
def board(tmp_path_factory):
    """A scratch database, built through the CLI exactly as a user would."""
    db = tmp_path_factory.mktemp("e2e") / "e2e.db"
    for args in FIXTURE:
        run_cli(db, list(args))
    # health normally arrives from workbook fill colour; set it directly here
    # so every state is represented without needing a styled xlsx fixture
    import sqlite3

    con = sqlite3.connect(db)
    for node_id, health in HEALTHS.items():
        con.execute("UPDATE nodes SET health = ? WHERE id = ?", (health, node_id))
    con.commit()
    con.close()
    return db


class _DistHandler(SimpleHTTPRequestHandler):
    def log_message(self, *args):  # keep pytest output clean
        pass

    def do_GET(self):
        if self.path == "/favicon.ico":  # never a console 404
            self.send_response(204)
            self.end_headers()
            return
        super().do_GET()


@pytest.fixture(scope="module")
def app_url():
    srv = ThreadingHTTPServer(
        ("127.0.0.1", 0), partial(_DistHandler, directory=str(DIST))
    )
    threading.Thread(target=srv.serve_forever, daemon=True).start()
    yield f"http://127.0.0.1:{srv.server_address[1]}/index.html"
    srv.shutdown()


@pytest.fixture(scope="module")
def browser():
    with sync_playwright() as pw:
        b = pw.chromium.launch()
        yield b
        b.close()


TAURI_STUB = """
    window.__TAURI__ = { core: { invoke: async (cmd, payload) => {
        if (cmd === "graph_html") return await window.__graphHtml();
        if (cmd === "env_info") return { bundled: false, working_dir: "e2e" };
        if (cmd !== "pt") throw new Error("unexpected command: " + cmd);
        return await window.__ptBridge(payload.args);
    } } };
"""
CONFIG_KEYS = """
    localStorage.setItem("dir", ".");
    localStorage.setItem("db", "e2e");
    localStorage.setItem("days", "30");
"""


def _wire(pg, db: Path, errors: list[str]) -> None:
    """Point the page's Tauri seam at the real CLI over a scratch db."""
    pg.on("pageerror", lambda e: errors.append(str(e)))
    pg.on(
        "console",
        lambda m: errors.append(m.text) if m.type == "error" else None,
    )
    pg.expose_function("__ptBridge", lambda args: run_cli(db, list(args)))

    def graph_html() -> str:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "g.html"
            run_cli(db, ["graph", str(out)])
            return out.read_text(encoding="utf8")

    pg.expose_function("__graphHtml", graph_html)


@pytest.fixture
def page(browser, app_url, board):
    """The dashboard, with its Tauri bridge pointed at the real CLI."""
    ctx = browser.new_context(viewport={"width": 1360, "height": 900})
    pg = ctx.new_page()
    errors: list[str] = []
    _wire(pg, board, errors)
    pg.add_init_script(TAURI_STUB + CONFIG_KEYS)
    pg.goto(app_url)
    pg.wait_for_selector('[data-testid="tree-node"]')
    pg.errors = errors
    yield pg
    ctx.close()


# --- small helpers ----------------------------------------------------------


def nav(page, view: str) -> None:
    """Everything lives on one page now; navigation is a no-op kept so the
    tests still read as 'go work in that panel' (Playwright scrolls to
    whatever it interacts with)."""


def toast(page):
    return page.locator('[data-testid="toast"]').first


def err_toast(page):
    return page.locator('[data-testid="toast"].err').first


def tree(page):
    return page.locator('[data-testid="tree"]')


def tree_nodes(page, kind: str | None = None):
    sel = '[data-testid="tree-node"]'
    if kind:
        sel += f'[data-kind="{kind}"]'
    return page.locator(sel)


def ready_rows(page):
    return page.locator('[data-testid="ready-row"]')


def make_workbook(path: Path, project: str, milestone: str, goal: str,
                  tasks: list[str], sheet: str | None = None) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet or project[:28]
    for col, head in enumerate(("Project", "Milestone", "Goal", "Task", "Notes"), 1):
        ws.cell(row=1, column=col, value=head)
    ws["A2"] = project
    ws["B3"] = milestone
    ws["C4"] = goal
    for i, task in enumerate(tasks):
        ws.cell(row=5 + i, column=4, value=task)
    wb.save(path)
    return path


def import_workbook(page, path: str) -> None:
    """Drive the three-step wizard from path to applied import."""
    page.locator('[data-testid="import-open"]').click()
    page.locator('[data-testid="import-path"]').fill(path)
    page.locator('[data-testid="import-next"]').click()
    expect(page.locator('[data-testid="import-preview-table"]')).to_be_visible()
    page.locator('[data-testid="import-apply"]').click()
    expect(page.locator('[data-testid="import-result"]')).to_contain_text("created")
    page.locator('[data-testid="import-done"]').click()


def graph_node(page, nid):
    return page.locator(f'[data-testid="graph-node"][data-nid="{nid}"]')


def fit_graph(page):
    """The panel opens 1:1 at the top; bring everything into the canvas so
    node coordinates are clickable."""
    page.locator('[data-testid="graph-fit"]').click()


def select_graph_node(page, nid):
    fit_graph(page)
    node = graph_node(page, nid)
    node.scroll_into_view_if_needed()
    node.click()
    expect(page.locator('[data-testid="graph-toolbar"]')).to_be_visible()


def drag_dependency(page, src: int, dst: int) -> None:
    """Draw src -> dst with the mouse: press the source card's port, release
    over the target card."""
    fit_graph(page)
    node = graph_node(page, src)
    node.scroll_into_view_if_needed()
    node.hover()  # the port fades in on hover
    port = node.locator('[data-testid="graph-port"]')
    pb = port.bounding_box()
    tb = graph_node(page, dst).bounding_box()
    page.mouse.move(pb["x"] + pb["width"] / 2, pb["y"] + pb["height"] / 2)
    page.mouse.down()
    page.mouse.move(tb["x"] + tb["width"] / 2, tb["y"] + tb["height"] / 2,
                    steps=8)
    page.mouse.up()


# --- the board screen -------------------------------------------------------


def test_tree_shows_the_whole_hierarchy(page):
    for name in ("Bridge Retrofit", "Construction", "Steelwork", "Survey the span"):
        expect(tree(page)).to_contain_text(name)
    # every kind is present
    kinds = {"project", "milestone", "goal", "task"}
    seen = set(tree_nodes(page).evaluate_all("els => els.map(e => e.dataset.kind)"))
    assert kinds <= seen, f"missing kinds: {kinds - seen}"


def test_ready_tasks_carry_a_green_circle(page, board):
    """The whole point of the dot: ready work is findable in the tree."""
    ready_names = {t["name"] for t in run_cli(board, ["ready"])}
    assert ready_names, "fixture must have ready work"
    rows = tree_nodes(page, "task")
    greens, labelled = set(), 0
    for i in range(rows.count()):
        row = rows.nth(i)
        dot = row.locator(".dot")
        colour = dot.evaluate("e => getComputedStyle(e).backgroundColor")
        if dot.get_attribute("title") == "ready":
            labelled += 1
            greens.add(row.locator(".label").inner_text())
            assert colour == GREEN, f"ready dot is {colour}, not green"
    assert greens == ready_names, f"green dots {greens} != ready set {ready_names}"
    assert labelled == len(ready_names)


def test_tree_collapses_and_expands(page):
    before = tree_nodes(page).count()
    page.locator('[data-testid="tree-toggle-all"]').click()  # collapse all
    assert tree_nodes(page).count() < before
    page.locator('[data-testid="tree-toggle-all"]').click()  # expand again
    assert tree_nodes(page).count() == before


def test_stat_row_reports_the_board(page):
    stats = page.locator(".stat")
    expect(stats).to_have_count(4)
    expect(page.locator('[data-testid="stat-projects"]')).to_contain_text("Projects")
    expect(page.locator('[data-testid="stat-ready"]')).to_contain_text("Ready now")
    # the tiles must agree with what they summarise, whatever the board holds
    shown = tree_nodes(page, "project").count()
    assert (page.locator('[data-testid="stat-projects"] .value').inner_text()
            == str(shown))
    assert (page.locator('[data-testid="stat-ready"] .value').inner_text()
            == str(ready_rows(page).count()))


def test_health_is_never_colour_alone(page):
    """Every health state must be named in text, not just painted."""
    legend = page.locator('[data-testid="legend"]')
    for label in ("on track", "not begun", "off track", "won't finish"):
        expect(legend).to_contain_text(label)
    # and the per-project breakdown carries counts beside each swatch
    expect(page.locator('[data-testid="tree"] .breakdown').first).to_be_visible()
    # the tree's own dots are named too, not just the health segments
    for label in ("ready", "blocked", "in progress", "waiting"):
        expect(legend).to_contain_text(label)


def test_health_bar_shows_a_segment_per_state(page):
    # Bridge Retrofit has three open tasks in three different health states,
    # so its stacked bar must render multiple visible segments.
    bar = page.locator('[data-testid="tree"] .bar').first
    expect(bar).to_be_visible()
    segments = bar.locator("span")
    assert segments.count() >= 2, "expected one bar segment per health state"
    expect(segments.first).to_be_visible()


def test_ready_list_is_ranked_by_impact(page, board):
    rows = ready_rows(page)
    assert rows.count() >= 2
    expect(rows.first).to_contain_text("unlocks")
    expect(rows.first).to_contain_text("gates")
    # ranking is decided by the command layer; the UI must not resort it
    expected = [t["name"] for t in run_cli(board, ["ready", "--impact"])]
    rendered = page.locator(
        '[data-testid="ready-row"] .title .name'
    ).all_inner_texts()
    assert rendered == expected


def test_selecting_a_project_filters_the_ready_list(page):
    before = ready_rows(page).count()
    project = tree_nodes(page, "project").filter(has_text="Bridge Retrofit")
    project.click()
    expect(page.locator('[data-testid="ready-scope"]')).to_have_text("Bridge Retrofit")
    assert ready_rows(page).count() < before
    # clicking again clears the filter
    project.click()
    expect(page.locator('[data-testid="ready-scope"]')).to_have_text("all projects")
    expect(ready_rows(page)).to_have_count(before)


def test_completing_a_task_reports_what_it_unlocked(page):
    first = ready_rows(page).first
    name = first.locator(".title .name").inner_text()
    first.locator('[data-testid="ready-tick"]').click()
    expect(toast(page)).to_contain_text(f"Completed “{name}”")
    expect(toast(page)).to_contain_text("Now ready")


def test_add_dialog_creates_a_project_through_the_cli(page):
    page.locator('[data-testid="add-open"]').click()
    page.locator('[data-testid="add-kind"]').select_option("project")
    page.locator('[data-testid="add-name"]').fill("Aqueduct")
    page.locator('[data-testid="add-ok"]').click()
    expect(toast(page)).to_contain_text("Created project")
    expect(tree(page)).to_contain_text("Aqueduct")


# --- the import wizard ------------------------------------------------------


def test_import_wizard_ingests_a_workbook(page, tmp_path):
    """The path most likely to be used first, and the one that has to work on
    a machine where the workbook does not sit next to the code."""
    book = make_workbook(
        tmp_path / "imported.xlsx", "Aqueduct Repair", "Phase 1", "Masonry",
        ["Point the arches"],
    )
    page.locator('[data-testid="import-open"]').click()
    page.locator('[data-testid="import-path"]').fill(str(book))  # absolute path
    page.locator('[data-testid="import-next"]').click()
    # step 2: the per-project plan, with nothing written yet
    row = page.locator('[data-testid="import-preview-row"]')
    expect(row).to_have_count(1)
    expect(row).to_contain_text("Aqueduct Repair")
    expect(row).to_contain_text("no match")
    page.locator('[data-testid="import-apply"]').click()
    expect(page.locator('[data-testid="import-result"]')).to_contain_text("created")
    page.locator('[data-testid="import-done"]').click()
    expect(tree(page)).to_contain_text("Aqueduct Repair")
    expect(tree(page)).to_contain_text("Point the arches")


def test_import_accepts_a_path_copied_from_explorer(page, tmp_path):
    """Explorer's "Copy as path" wraps the path in quotes and argv carries
    them through, so the extension check used to fail on `.xlsx"`."""
    book = make_workbook(
        tmp_path / "quoted.xlsx", "Quoted Path Project", "Phase", "Goal",
        ["A task"], sheet="Quoted",
    )
    import_workbook(page, f'"{book}"')  # exactly what is pasted
    expect(tree(page)).to_contain_text("Quoted Path Project")


def test_import_of_an_unreadable_file_reports_a_usable_error(page, tmp_path):
    junk = tmp_path / "notes.txt"
    junk.write_text("not a workbook", encoding="utf8")
    page.locator('[data-testid="import-open"]').click()
    page.locator('[data-testid="import-path"]').fill(str(junk))
    page.locator('[data-testid="import-next"]').click()
    t = err_toast(page)
    expect(t).to_be_visible()
    # a real reason, not a traceback and not {"name":"Error"}
    expect(t).to_contain_text("could not read")
    assert "Traceback" not in t.inner_text()
    # the wizard stays on step 1 so the path can be corrected
    expect(page.locator('[data-testid="import-path"]')).to_be_visible()


def test_import_of_a_missing_workbook_reports_a_usable_error(page):
    page.locator('[data-testid="import-open"]').click()
    page.locator('[data-testid="import-path"]').fill("no-such-workbook.xlsx")
    page.locator('[data-testid="import-next"]').click()
    t = err_toast(page)
    expect(t).to_be_visible()
    expect(t).to_contain_text("workbook not found")
    # the wizard must stay open so the path can be corrected
    expect(page.locator('[data-testid="import-path"]')).to_be_visible()


def test_no_console_or_page_errors(page):
    page.locator('[data-testid="refresh"]').click()
    expect(page.locator('[data-testid="refresh"]')).to_be_enabled()
    assert page.errors == [], f"page reported errors: {page.errors}"


# --- editing, journal, priorities ------------------------------------------


def test_edit_dialog_renames_through_the_cli(page):
    row = tree_nodes(page, "goal").first
    row.hover()
    row.locator('[data-testid="tree-edit"]').click()
    dlg = page.locator('[data-testid="edit-dialog"]')
    expect(dlg).to_be_visible()
    page.locator('[data-testid="edit-name"]').fill("Steelwork (revised)")
    page.locator('[data-testid="edit-ok"]').click()
    expect(toast(page)).to_contain_text("Saved")
    expect(tree(page)).to_contain_text("Steelwork (revised)")


def test_edit_dialog_adds_and_removes_a_dependency(page, board):
    # goal Optics depends on goal Steelwork* -> Lens spec becomes blocked;
    # removing the edge in the same dialog makes it ready again.
    optics = next(n for n in run_cli(board, ["ls"]) if n["name"] == "Optics")
    row = tree_nodes(page, "goal").filter(has_text="Optics")
    row.hover()
    row.locator('[data-testid="tree-edit"]').click()
    dlg = page.locator('[data-testid="edit-dialog"]')
    expect(dlg).to_be_visible()
    pick = page.locator('[data-testid="edit-dep-pick"]')
    steel = pick.locator("option", has_text="Steelwork").first.get_attribute("value")
    pick.select_option(steel)
    page.locator('[data-testid="edit-dep-add"]').click()
    expect(toast(page)).to_contain_text("Dependency added")
    deps = run_cli(board, ["dep", "ls"])
    assert any(
        d["from_id"] == int(steel) and d["to_id"] == optics["id"] for d in deps
    )
    # visible consequence: Lens spec is no longer ready
    assert "Lens spec" not in {t["name"] for t in run_cli(board, ["ready"])}
    # now remove it through the same dialog
    dep_row = dlg.locator(".dep-row", has_text="Steelwork")
    expect(dep_row).to_be_visible()
    dep_row.locator("button.x").click()
    expect(dlg.locator(".dep-row", has_text="Steelwork")).to_have_count(0)
    deps = run_cli(board, ["dep", "ls"])
    assert not any(
        d["from_id"] == int(steel) and d["to_id"] == optics["id"] for d in deps
    )
    assert "Lens spec" in {t["name"] for t in run_cli(board, ["ready"])}
    dlg.locator("button", has_text="Cancel").click()


def test_capture_appears_in_the_journal(page):
    nav(page, "journal")
    page.locator('[data-testid="capture-input"]').fill(
        "tried the elastic modulus rig"
    )
    page.locator('[data-testid="capture-submit"]').click()
    expect(page.locator('[data-testid="journal-days"]')).to_contain_text(
        "tried the elastic modulus rig"
    )


def test_pinned_task_leads_the_todo_list(page, board):
    tid = ready_rows(page).last.get_attribute("data-nid")
    run_cli(board, ["set", tid, "--priority", "pinned"])
    try:
        page.locator('[data-testid="refresh"]').click()
        first = ready_rows(page).first
        expect(first.locator(".prio")).to_have_text("pinned")
        assert first.get_attribute("data-nid") == tid
    finally:
        run_cli(board, ["set", tid, "--priority", "none"])


# --- the native graph editor: draw, provenance, and verbs on the canvas -----


def test_graph_panel_renders_every_open_task(page, board):
    """One card per open task, straight from graph-data; done tasks hidden
    until the toggle asks for them."""
    panel = page.locator('[data-testid="graph-panel"]')
    expect(panel).to_be_visible()
    data = run_cli(board, ["graph-data"])
    open_tasks = [
        n for n in data["nodes"]
        if n["kind"] == "task"
        and n.get("state") not in ("done", "dropped")
    ]
    assert open_tasks, "fixture must have open tasks"
    for n in open_tasks:
        expect(graph_node(page, n["id"])).to_be_visible()
    done = [
        n for n in data["nodes"]
        if n["kind"] == "task" and n.get("state") == "done"
    ]
    for n in done:
        expect(graph_node(page, n["id"])).to_have_count(0)


def test_graph_provenance_is_line_style_not_colour_alone(page, board):
    """The point of the editor: an entry-order guess renders as a different
    edge kind than a rank the user chose, and an overruled guess is still
    there, ghosted — plus a legend naming all of it."""
    data = run_cli(board, ["graph-data"])
    kinds = {e["kind"] for e in data["edges"]}
    assert "seq-assumed" in kinds, "fixture tasks are auto-ranked guesses"
    edges = page.locator('[data-testid="graph-edge"]')
    assert edges.count() > 0
    # (edges can be perfectly horizontal, giving a zero-height bounding box,
    # so "attached" is the right assertion for them, not "visible")
    assumed = page.locator('[data-testid="graph-edge"][data-kind="seq-assumed"]')
    expect(assumed.first).to_be_attached()
    legend = page.locator('[data-testid="graph-panel"] .legend')
    for label in ("explicit dependency", "chosen order", "guessed order",
                  "overruled guess"):
        expect(legend).to_contain_text(label)
    # the toggle removes guessed edges from the drawing entirely
    page.locator('[data-testid="graph-show-guessed"]').uncheck()
    expect(
        page.locator('[data-testid="graph-edge"][data-kind="seq-assumed"]')
    ).to_have_count(0)
    page.locator('[data-testid="graph-show-guessed"]').check()
    expect(assumed.first).to_be_attached()


def test_graph_check_off_completes_through_the_bridge(page, board):
    ready = run_cli(board, ["ready"])[0]
    select_graph_node(page, ready["id"])
    page.locator('[data-testid="graph-node-done"]').click()
    expect(toast(page)).to_contain_text("Completed")
    node = next(n for n in run_cli(board, ["ls"]) if n["id"] == ready["id"])
    assert node["status"] == "done", "the graph tick must run the done verb"


def test_graph_start_and_pause_from_the_toolbar(page, board):
    ready = run_cli(board, ["ready"])[0]
    select_graph_node(page, ready["id"])
    page.locator('[data-testid="graph-node-start"]').click()
    expect(toast(page)).to_contain_text("Started")
    node = next(n for n in run_cli(board, ["ls"]) if n["id"] == ready["id"])
    assert node["status"] == "in_progress"
    expect(graph_node(page, ready["id"])).to_have_attribute(
        "data-state", "in_progress"
    )
    select_graph_node(page, ready["id"])
    page.locator('[data-testid="graph-node-pause"]').click()
    expect(toast(page)).to_contain_text("Paused")
    node = next(n for n in run_cli(board, ["ls"]) if n["id"] == ready["id"])
    assert node["status"] == "active", "pause must return the stored status"


def test_graph_edit_button_opens_the_app_edit_dialog(page, board):
    ready = run_cli(board, ["ready"])[0]
    select_graph_node(page, ready["id"])
    page.locator('[data-testid="graph-node-edit"]').click()
    dlg = page.locator('[data-testid="edit-dialog"]')
    expect(dlg).to_be_visible()
    dlg.locator("button", has_text="Cancel").click()


def test_graph_draws_a_dependency_and_rejects_cycles_gracefully(page, board):
    """Drag the port from one ready task onto another: dep add through the
    same pt door. The reverse drag must be rejected with the cycle path, and
    the removal path (select the edge, Remove) restores the board."""
    import itertools

    ready = run_cli(board, ["ready"])
    assert len(ready) >= 2, "need two ready tasks to connect"
    existing = {(d["from_id"], d["to_id"])
                for d in run_cli(board, ["dep", "ls"])}
    src, dst = next(
        (a, b)
        for a, b in itertools.permutations([t["id"] for t in ready], 2)
        if (a, b) not in existing and (b, a) not in existing
    )
    before = len(run_cli(board, ["dep", "ls"]))

    drag_dependency(page, src, dst)
    expect(toast(page)).to_contain_text("Dependency added")
    deps = run_cli(board, ["dep", "ls"])
    assert len(deps) == before + 1
    assert any(d["from_id"] == src and d["to_id"] == dst for d in deps)
    edge = page.locator(
        f'[data-testid="graph-edge"][data-kind="dep"]'
        f'[data-from="{src}"][data-to="{dst}"]'
    )
    expect(edge).to_be_attached()

    # the reverse edge closes a loop: structured rejection, nothing written
    drag_dependency(page, dst, src)
    t = err_toast(page)
    expect(t).to_be_visible()
    expect(t).to_contain_text("rejected")
    assert len(run_cli(board, ["dep", "ls"])) == before + 1

    # remove it on the canvas: click the edge's midpoint (a zero-height
    # horizontal path can't be clicked as an element), then Remove
    page.locator('[data-testid="graph-panel"] svg').scroll_into_view_if_needed()
    mid = edge.locator(".line").evaluate(
        """(p) => {
            const m = p.getPointAtLength(p.getTotalLength() / 2);
            const s = new DOMPoint(m.x, m.y)
                .matrixTransform(p.ownerSVGElement.getScreenCTM());
            return {x: s.x, y: s.y};
        }"""
    )
    page.mouse.click(mid["x"], mid["y"])
    page.locator('[data-testid="graph-edge-rm"]').click()
    # (earlier toasts from this same test still sit in the stack, so the
    # proof of removal is the drawing and the database, not toast .first)
    expect(edge).to_have_count(0)
    assert len(run_cli(board, ["dep", "ls"])) == before


def test_graph_rank_buttons_author_a_user_rank(page, board):
    """Mouse-only rank authoring (spec §11.5): nudging a guessed rank turns
    it into a chosen one, and the edge kind follows."""
    gid = run_cli(board, ["add", "goal", "Ranking", "--parent", "2"])
    gid = gid["created"]["id"]
    a = run_cli(board, ["add", "task", "rank a", "--parent", str(gid)])
    b = run_cli(board, ["add", "task", "rank b", "--parent", str(gid)])
    a, b = a["created"]["id"], b["created"]["id"]
    page.locator('[data-testid="refresh"]').click()

    select_graph_node(page, b)
    toolbar = page.locator('[data-testid="graph-toolbar"]')
    expect(toolbar).to_contain_text("rank 2")
    expect(toolbar).to_contain_text("(guessed)")
    page.locator('[data-testid="graph-rank-down"]').click()
    # equal ranks are parallel: the a->b guess leaves the drawing, which is
    # also the signal that the panel refetched after the verb
    expect(
        page.locator(f'[data-testid="graph-edge"][data-from="{a}"][data-to="{b}"]')
    ).to_have_count(0)

    node_b = next(n for n in run_cli(board, ["ls"]) if n["id"] == b)
    assert node_b["seq_index"] == 1
    assert node_b["seq_source"] == "user", "a clicked rank is user provenance"
    select_graph_node(page, b)
    expect(page.locator('[data-testid="graph-toolbar"]')).not_to_contain_text(
        "(guessed)"
    )


def test_graph_wait_badge_carries_date_and_reason(page, board):
    tid = run_cli(board, ["add", "task", "Await castings"])["created"]["id"]
    run_cli(board, ["wait", str(tid), "--until", "2027-02-01",
                    "--reason", "foundry lead time"])
    page.locator('[data-testid="refresh"]').click()
    node = graph_node(page, tid)
    expect(node).to_have_attribute("data-state", "waiting")
    expect(node).to_contain_text("⏳ 2027-02-01")
    run_cli(board, ["rm", str(tid), "--yes"])  # leave the board as found
    page.locator('[data-testid="refresh"]').click()


# --- first run --------------------------------------------------------------


def test_setup_first_run_starts_empty(browser, app_url, tmp_path):
    """No `db` key in localStorage -> the setup screen; "start empty" writes
    the config and lands on the Board."""
    db = tmp_path / "fresh.db"
    ctx = browser.new_context(viewport={"width": 1360, "height": 900})
    pg = ctx.new_page()
    errors: list[str] = []
    _wire(pg, db, errors)
    pg.add_init_script(TAURI_STUB)  # note: no localStorage config at all
    pg.goto(app_url)
    expect(pg.locator('[data-testid="setup-screen"]')).to_be_visible()
    pg.locator('[data-testid="setup-start-empty"]').click()
    # the board renders (its stat tiles are unconditional) ...
    expect(pg.locator(".stat")).to_have_count(4)
    expect(pg.locator('[data-testid="setup-screen"]')).to_have_count(0)
    # ... and the config is persisted so setup never reappears
    assert pg.evaluate("localStorage.getItem('db')") == "tracker.db"
    assert errors == [], f"page reported errors: {errors}"
    ctx.close()


# --- the Today screen -------------------------------------------------------


def today_rows(page):
    return page.locator('[data-testid="today-row"]')


def quick_add(page, name: str) -> None:
    page.locator('[data-testid="today-quick-add"]').fill(name)
    page.locator('[data-testid="today-quick-add-ok"]').click()
    expect(page.locator('[data-testid="today-list"]')).to_contain_text(name)


def test_today_quick_add_creates_and_lists(page, board):
    nav(page, "today")
    quick_add(page, "Sand the skirting boards")
    expect(toast(page)).to_contain_text("Added “Sand the skirting boards”")
    # the quick-added item really is a task in the database
    names = {n["name"] for n in run_cli(board, ["ls"])}
    assert "Sand the skirting boards" in names


def test_today_pulls_a_ready_task_onto_the_list(page):
    nav(page, "today")
    page.locator('[data-testid="today-pool-toggle"]').click()
    pool_row = page.locator('[data-testid="today-pool-row"]').first
    name = pool_row.locator(".name").first.inner_text()
    pool_row.locator('[data-testid="today-pool-add"]').click()
    listed = today_rows(page).filter(has_text=name)
    expect(listed.first).to_be_visible()


def test_today_tick_moves_the_item_to_completed(page):
    nav(page, "today")
    quick_add(page, "Grease the winch")
    row = today_rows(page).filter(has_text="Grease the winch")
    row.locator('[data-testid="today-tick"]').click()
    expect(toast(page)).to_contain_text("Completed “Grease the winch”")
    expect(page.locator('[data-testid="today-completed"]')).to_contain_text(
        "Grease the winch"
    )
    expect(today_rows(page).filter(has_text="Grease the winch")).to_have_count(0)


def test_today_remove_tombstones_the_item(page):
    nav(page, "today")
    quick_add(page, "Coil the ropes")
    row = today_rows(page).filter(has_text="Coil the ropes")
    row.hover()
    row.locator('[data-testid="today-remove"]').click()
    expect(today_rows(page).filter(has_text="Coil the ropes")).to_have_count(0)
    # tombstoned, not resurrected: a refresh does not bring it back
    page.locator('[data-testid="refresh"]').click()
    expect(page.locator('[data-testid="refresh"]')).to_be_enabled()
    expect(today_rows(page).filter(has_text="Coil the ropes")).to_have_count(0)


def test_today_move_buttons_reorder_the_list(page):
    nav(page, "today")
    quick_add(page, "Alpha step")
    quick_add(page, "Beta step")
    rows = today_rows(page)
    n = rows.count()
    assert n >= 2
    # the two quick-adds land last: [..., Alpha step, Beta step]
    expect(rows.nth(n - 1).locator(".name").first).to_have_text("Beta step")
    last = rows.nth(n - 1)
    last.hover()
    last.locator('button[title="Move up"]').click()
    expect(today_rows(page).nth(n - 2).locator(".name").first).to_have_text(
        "Beta step"
    )
    expect(today_rows(page).nth(n - 1).locator(".name").first).to_have_text(
        "Alpha step"
    )


# --- reminders / Upcoming ---------------------------------------------------


def test_reminder_planned_for_tomorrow_shows_in_upcoming(page, board):
    # (auto-landing on Today when the day arrives needs an injected clock,
    # which the UI cannot fake — that path is unit-tested in test_today.py)
    tomorrow = (
        date.fromisoformat(run_cli(board, ["today"])["date"]) + timedelta(days=1)
    ).isoformat()
    nav(page, "upcoming")
    page.locator('[data-testid="remind-name"]').fill("Renew the crane permit")
    page.locator('[data-testid="remind-in-days"]').fill("1")
    page.locator('[data-testid="remind-ok"]').click()
    expect(toast(page)).to_contain_text("Planned “Renew the crane permit”")
    expect(toast(page)).to_contain_text(tomorrow)
    row = page.locator('[data-testid="upcoming-row"]', has_text="Renew the crane permit")
    expect(row).to_be_visible()
    group = page.locator(".up-group", has=row)
    expect(group.locator(".up-date")).to_contain_text(tomorrow)


def test_recurring_reminder_and_wait_reason_render(page, board):
    """P2 surfaces: a reminder planned with a repeat rule carries the
    recurring badge, and a task parked with `wait --reason` shows what it
    waits on (both straight from the command layer's data)."""
    nav(page, "upcoming")
    page.locator('[data-testid="remind-name"]').fill("Log loop pressures")
    page.locator('[data-testid="remind-in-days"]').fill("2")
    page.locator('[data-testid="remind-every"]').select_option("weekly")
    page.locator('[data-testid="remind-ok"]').click()
    expect(toast(page)).to_contain_text("repeating weekly")
    row = page.locator('[data-testid="upcoming-row"]',
                       has_text="Log loop pressures")
    expect(row.locator('[data-testid="badge-recurring"]')).to_be_visible()

    # the wait verb is CLI-first; the UI must render its reason wherever
    # the task appears
    tid = run_cli(board, ["add", "task", "Await pump delivery"])["created"]["id"]
    run_cli(board, ["wait", str(tid), "--until", "2027-01-05",
                    "--reason", "3-week vendor lead time"])
    page.locator('[data-testid="refresh"]').click()
    wrow = page.locator('[data-testid="upcoming-row"]',
                        has_text="Await pump delivery")
    expect(wrow.locator('[data-testid="badge-wait-reason"]')).to_contain_text(
        "3-week vendor lead time"
    )


# --- import wizard: the name-collision choice -------------------------------


def test_import_same_name_defaults_to_a_new_project(page, tmp_path):
    """A bare name coincidence must not silently merge two trackers: the
    preview surfaces the match and suggests "new", and applying it leaves two
    distinct projects in the tree."""
    book_a = make_workbook(
        tmp_path / "depot_a.xlsx", "Twin Peaks Depot", "Opening", "Stock",
        ["Stock the shelves"], sheet="DepotA",
    )
    book_b = make_workbook(
        tmp_path / "depot_b.xlsx", "Twin Peaks Depot", "Rebrand", "Frontage",
        ["Repaint the sign"], sheet="DepotB",
    )
    import_workbook(page, str(book_a))
    expect(tree(page)).to_contain_text("Twin Peaks Depot")

    page.locator('[data-testid="import-open"]').click()
    page.locator('[data-testid="import-path"]').fill(str(book_b))
    page.locator('[data-testid="import-next"]').click()
    row = page.locator('[data-testid="import-preview-row"]')
    expect(row).to_contain_text("matched by name")
    choice = row.locator('[data-testid="import-project-choice"]')
    assert choice.input_value() == "new", "a name match must default to 'new'"
    page.locator('[data-testid="import-apply"]').click()
    expect(page.locator('[data-testid="import-result"]')).to_contain_text("created")
    page.locator('[data-testid="import-done"]').click()

    twins = tree_nodes(page, "project").filter(has_text="Twin Peaks Depot")
    expect(twins).to_have_count(2)
    expect(tree(page)).to_contain_text("Stock the shelves")
    expect(tree(page)).to_contain_text("Repaint the sign")


# --- per-task notes ---------------------------------------------------------


def test_task_notes_tab_and_journal_search(page):
    row = tree_nodes(page, "task").filter(has_text="Erect falsework")
    row.hover()
    row.locator('[data-testid="tree-edit"]').click()
    dlg = page.locator('[data-testid="edit-dialog"]')
    expect(dlg).to_be_visible()
    page.locator('[data-testid="notes-tab"]').click()
    page.locator('[data-testid="note-input"]').fill("checked the plumb lines twice")
    page.locator('[data-testid="note-add"]').click()
    note = dlg.locator('[data-testid="note-row"]', has_text="plumb lines")
    expect(note).to_be_visible()
    dlg.locator("button", has_text="Close").click()

    nav(page, "journal")
    page.locator('[data-testid="journal-search"]').fill("plumb lines")
    page.locator("button", has_text="Search").click()
    result = page.locator(
        '[data-testid="journal-results"] [data-testid="note-row"]',
        has_text="plumb lines",
    )
    expect(result).to_be_visible()
    expect(result.locator(".node-chip")).to_have_text("Erect falsework")


# --- more graph editor: bands, edge endpoints, suppression -------------------


def test_graph_groups_planner_tasks_in_their_own_band(page, board):
    """The fixture's planner task (no project) renders under a 'planner'
    band label, projects under theirs."""
    panel = page.locator('[data-testid="graph-panel"]')
    expect(panel.locator(".g-band", has_text="planner")).to_be_visible()
    expect(panel.locator(".g-band", has_text="Bridge Retrofit")).to_be_visible()
    rig = next(n for n in run_cli(board, ["ls"])
               if n["name"] == "Calibrate the strain rig")
    expect(graph_node(page, rig["id"])).to_be_visible()


def test_graph_edges_start_and_end_on_their_cards(page, board):
    """Native SVG: every drawn edge's endpoints must sit on the borders of
    the cards it links — the alignment bug class the old iframe had."""
    geo = page.locator('[data-testid="graph-panel"] svg').evaluate(
        """(svg) => {
            const out = [];
            for (const g of svg.querySelectorAll('[data-testid="graph-edge"]')) {
              const line = g.querySelector('.line');
              const p0 = line.getPointAtLength(0);
              const p1 = line.getPointAtLength(line.getTotalLength());
              const find = id =>
                svg.querySelector(`[data-testid="graph-node"][data-nid="${id}"]`)
                   ?.getAttribute('transform');
              out.push({from: g.dataset.from, to: g.dataset.to,
                        x0: p0.x, y0: p0.y, x1: p1.x, y1: p1.y,
                        tFrom: find(g.dataset.from), tTo: find(g.dataset.to)});
            }
            return out;
        }"""
    )
    assert geo, "the panel must draw edges"
    for e in geo:
        assert e["tFrom"] and e["tTo"], f"edge {e} links a missing card"
        fx, fy = map(float, e["tFrom"][10:-1].split())
        tx, ty = map(float, e["tTo"][10:-1].split())
        # starts on the source's right edge, ends on the target's left edge
        assert abs(e["x0"] - (fx + 188)) < 1 and abs(e["x1"] - tx) < 1
        assert fy <= e["y0"] <= fy + 56 and ty <= e["y1"] <= ty + 56


def test_graph_ghosts_a_suppressed_guess(page, board):
    """Draw an explicit edge into a task with a guessed rank: the guess is
    overruled and the drawing shows it dotted-and-faint, not gone."""
    gid = run_cli(board, ["add", "goal", "Suppression", "--parent", "2"])
    gid = gid["created"]["id"]
    t1 = run_cli(board, ["add", "task", "sup 1", "--parent", str(gid)])
    t2 = run_cli(board, ["add", "task", "sup 2", "--parent", str(gid)])
    other = run_cli(board, ["add", "task", "sup outsider"])
    t1, t2 = t1["created"]["id"], t2["created"]["id"]
    other = other["created"]["id"]
    run_cli(board, ["dep", "add", str(other), str(t2)])
    page.locator('[data-testid="refresh"]').click()

    ghost = page.locator(
        f'[data-testid="graph-edge"][data-kind="seq-suppressed"]'
        f'[data-from="{t1}"][data-to="{t2}"]'
    )
    expect(ghost).to_be_attached()
    dep = page.locator(
        f'[data-testid="graph-edge"][data-kind="dep"]'
        f'[data-from="{other}"][data-to="{t2}"]'
    )
    expect(dep).to_be_attached()


# --- in progress: start and pause from the lists -----------------------------


def test_ready_row_start_moves_the_task_to_in_progress(page, board):
    row = ready_rows(page).first
    nid = row.get_attribute("data-nid")
    name = row.locator(".title .name").inner_text()
    row.hover()
    row.locator('[data-testid="ready-start"]').click()
    expect(toast(page)).to_contain_text(f"Started “{name}”")
    # the task leaves the ready list and appears in the In progress strip
    strip = page.locator(f'[data-testid="inprogress-row"][data-nid="{nid}"]')
    expect(strip).to_be_visible()
    expect(strip).to_contain_text("in progress")
    expect(
        page.locator(f'[data-testid="ready-row"][data-nid="{nid}"]')
    ).to_have_count(0)
    node = next(n for n in run_cli(board, ["ls"]) if n["id"] == int(nid))
    assert node["status"] == "in_progress"

    # pause sends it straight back
    strip.hover()
    strip.locator('[data-testid="inprogress-pause"]').click()
    expect(toast(page)).to_contain_text(f"Paused “{name}”")
    expect(
        page.locator(f'[data-testid="ready-row"][data-nid="{nid}"]')
    ).to_be_visible()
    expect(
        page.locator(f'[data-testid="inprogress-row"][data-nid="{nid}"]')
    ).to_have_count(0)


def test_in_progress_strip_tick_completes_the_task(page, board):
    row = ready_rows(page).first
    nid = row.get_attribute("data-nid")
    row.hover()
    row.locator('[data-testid="ready-start"]').click()
    strip = page.locator(f'[data-testid="inprogress-row"][data-nid="{nid}"]')
    expect(strip).to_be_visible()
    strip.locator('[data-testid="inprogress-tick"]').click()
    expect(toast(page)).to_contain_text("Completed")
    node = next(n for n in run_cli(board, ["ls"]) if n["id"] == int(nid))
    assert node["status"] == "done"


def test_today_row_start_shows_the_badge(page, board):
    nav(page, "today")
    quick_add(page, "Deburr the fixture plates")
    row = today_rows(page).filter(has_text="Deburr the fixture plates")
    row.hover()
    row.locator('[data-testid="today-start"]').click()
    expect(toast(page)).to_contain_text("Started")
    expect(row.locator('[data-testid="badge-in-progress"]')).to_be_visible()
    # and pausing from the same row clears it
    row.hover()
    row.locator('[data-testid="today-pause"]').click()
    expect(toast(page)).to_contain_text("Paused")
    expect(row.locator('[data-testid="badge-in-progress"]')).to_have_count(0)


# --- P3: steps, find, suggestions, hashtag quick-add ------------------------


def test_steps_checklist_in_the_edit_dialog(page, board):
    """Steps live in the task drawer: add two, tick one, and the checklist
    state is the command layer's — visible again on the Today badge."""
    row = tree_nodes(page, "task").filter(has_text="Survey the span").first
    row.hover()
    row.locator('[data-testid="tree-edit"]').click()
    expect(page.locator('[data-testid="edit-dialog"]')).to_be_visible()
    page.locator('[data-testid="steps-tab"]').click()
    for name in ("stake the corners", "shoot elevations"):
        page.locator('[data-testid="step-input"]').fill(name)
        page.locator('[data-testid="step-add"]').click()
    rows = page.locator('[data-testid="step-row"]')
    expect(rows).to_have_count(2)
    rows.first.locator('[data-testid="step-check"]').click()
    expect(page.locator('[data-testid="steps-tab"]')).to_contain_text("1/2")
    expect(rows.first.locator('[data-testid="step-check"]')).to_be_checked()
    # the database agrees
    survey = next(n for n in run_cli(board, ["ls"])
                  if n["name"] == "Survey the span")
    r = run_cli(board, ["step", "ls", str(survey["id"])])
    assert (r["done"], r["total"]) == (1, 2)
    page.keyboard.press("Escape")


def test_find_modal_searches_and_opens_the_editor(page, board):
    page.keyboard.press("Control+k")
    expect(page.locator('[data-testid="find-modal"]')).to_be_visible()
    page.locator('[data-testid="find-input"]').fill("falsework")
    hit = page.locator('[data-testid="find-node"]', has_text="Erect falsework")
    expect(hit).to_be_visible()
    hit.click()
    dlg = page.locator('[data-testid="edit-dialog"]')
    expect(dlg).to_be_visible()
    expect(dlg).to_contain_text("Edit task")
    expect(page.locator('[data-testid="edit-name"]')).to_have_value(
        "Erect falsework"
    )
    page.keyboard.press("Escape")


def test_quick_add_hashtags_become_tags(page, board):
    nav(page, "today")
    # not the quick_add helper: the hashtags are deliberately NOT part of
    # the created name, so the list shows the stripped form
    page.locator('[data-testid="today-quick-add"]').fill(
        "swap quench bath fluid #lab #maintenance"
    )
    page.locator('[data-testid="today-quick-add-ok"]').click()
    expect(toast(page)).to_contain_text("Added “swap quench bath fluid”")
    expect(
        today_rows(page).filter(has_text="swap quench bath fluid").first
    ).to_be_visible()
    made = next(n for n in run_cli(board, ["ls"])
                if n["name"] == "swap quench bath fluid")
    assert made["tags"] == ["lab", "maintenance"]


def test_suggestions_offer_ready_work_with_reasons(page):
    nav(page, "today")
    sug = page.locator('[data-testid="today-suggestion"]')
    expect(sug.first).to_be_visible()
    name = sug.first.locator(".name").first.inner_text()
    sug.first.locator('[data-testid="today-suggestion-add"]').click()
    listed = today_rows(page).filter(has_text=name)
    expect(listed.first).to_be_visible()


# --- export / re-import round trip ------------------------------------------
#
# remind, priority, followup_days, earliest_start and Today membership now
# have columns in the export. These drive the full loop through the real UI:
# a board seeded over the CLI, exported, imported into a fresh database via
# the import wizard, and verified screen by screen — then re-imported to
# prove idempotence, and cycled export->import->export to prove stability.


def _fresh_page(browser, app_url, db):
    """A dashboard context of its own, bound to `db` (module page shares
    state across tests; round-trip tests need a deterministic board)."""
    ctx = browser.new_context(viewport={"width": 1360, "height": 900})
    pg = ctx.new_page()
    errors: list[str] = []
    _wire(pg, db, errors)
    pg.add_init_script(TAURI_STUB + CONFIG_KEYS)
    pg.goto(app_url)
    pg.errors = errors
    return ctx, pg


def _seed_rich_board(db: Path) -> None:
    """Every planner field in play, exactly as a user would set them."""
    for args in [
        ("add", "project", "Dyno Cell"),
        ("add", "milestone", "Commissioning", "--parent", "1"),
        ("add", "goal", "Sensors", "--parent", "2"),
        ("add", "task", "Mount load cell", "--parent", "3",
         "--priority", "high", "--followup-days", "3"),
        ("add", "task", "Wire thermocouples", "--parent", "3",
         "--priority", "low"),
        ("add", "task", "Order spare fuses", "--parent", "3"),
        ("today", "add", "5"),
        ("today", "add", "4"),
        ("remind", "--new", "Check calibration drift", "--in", "7"),
    ]:
        run_cli(db, list(args))


def test_export_import_round_trip_survives_the_real_ui(
    browser, app_url, tmp_path
):
    src = tmp_path / "src.db"
    _seed_rich_board(src)
    book = tmp_path / "handoff.xlsx"
    run_cli(src, ["export", str(book)])

    dst = tmp_path / "dst.db"
    ctx, pg = _fresh_page(browser, app_url, dst)
    try:
        expect(pg.locator(".stat")).to_have_count(4)
        import_workbook(pg, str(book))
        expect(tree(pg)).to_contain_text("Dyno Cell")

        # Today survived with its order (Wire thermocouples was added first)
        nav(pg, "today")
        expect(today_rows(pg)).to_have_count(2)
        expect(today_rows(pg).nth(0)).to_contain_text("Wire thermocouples")
        expect(today_rows(pg).nth(1)).to_contain_text("Mount load cell")

        # the reminder is still waiting on its day
        nav(pg, "upcoming")
        expect(
            pg.locator('[data-testid="upcoming-row"]',
                       has_text="Check calibration drift")
        ).to_be_visible()

        # priority survived and renders: the ready task carries its badge
        # (the other two are chained behind it, so only this one is ready)
        nav(pg, "board")
        row = ready_rows(pg).filter(has_text="Mount load cell")
        expect(row.locator(".prio")).to_have_text("high")

        # and the database agrees cell for cell
        by_name = {n["name"]: n for n in run_cli(dst, ["ls"])}
        assert by_name["Mount load cell"]["priority"] == "high"
        assert by_name["Mount load cell"]["followup_days"] == 3
        assert by_name["Wire thermocouples"]["priority"] == "low"
        assert by_name["Check calibration drift"]["remind"] == 1
        assert by_name["Check calibration drift"]["earliest_start"] is not None
        assert pg.errors == [], f"page reported errors: {pg.errors}"
    finally:
        ctx.close()


def test_reimport_through_the_wizard_is_idempotent(browser, app_url, tmp_path):
    """Importing our own export twice adds nothing: the wizard suggests
    merge (ref identity), and the Today list neither duplicates nor
    resurrects anything — including a task dismissed in the meantime."""
    src = tmp_path / "src.db"
    _seed_rich_board(src)
    book = tmp_path / "handoff.xlsx"
    run_cli(src, ["export", str(book)])

    dst = tmp_path / "dst.db"
    ctx, pg = _fresh_page(browser, app_url, dst)
    try:
        import_workbook(pg, str(book))
        before = len(run_cli(dst, ["ls"]))

        nav(pg, "today")
        row = today_rows(pg).filter(has_text="Mount load cell")
        row.hover()
        row.locator('[data-testid="today-remove"]').click()
        expect(today_rows(pg)).to_have_count(1)

        # second pass: preview must suggest merge, not a duplicate project
        pg.locator('[data-testid="import-open"]').click()
        pg.locator('[data-testid="import-path"]').fill(str(book))
        pg.locator('[data-testid="import-next"]').click()
        choice = pg.locator('[data-testid="import-project-choice"]').first
        expect(choice).to_have_value("merge")
        pg.locator('[data-testid="import-apply"]').click()
        expect(pg.locator('[data-testid="import-result"]')).to_be_visible()
        pg.locator('[data-testid="import-done"]').click()

        assert len(run_cli(dst, ["ls"])) == before  # not one node duplicated
        nav(pg, "today")
        expect(today_rows(pg)).to_have_count(1)  # tombstone outranked the file
        expect(today_rows(pg).first).to_contain_text("Wire thermocouples")
        assert pg.errors == [], f"page reported errors: {pg.errors}"
    finally:
        ctx.close()


def test_round_trip_reaches_a_fixed_point_over_three_cycles(tmp_path):
    """export -> import -> export ... must stabilise: after the first cycle
    every later workbook carries identical cell content, and the database
    stops changing. Guards against drift bugs (a column that re-imports
    slightly differently and mutates on every sync)."""
    db = tmp_path / "cycle.db"
    _seed_rich_board(db)

    def snapshot(d):
        nodes = run_cli(d, ["ls"])
        keep = ("name", "priority", "followup_days", "remind",
                "earliest_start", "deadline", "est_minutes", "ref", "status")
        return sorted(tuple(n.get(k) for k in keep) for n in nodes)

    def cells(path):
        from protracker.importer import read_workbook_rows

        return {
            title: [(r[0], r[1]) for r in rows]
            for title, rows in read_workbook_rows(str(path)).items()
        }

    books = []
    for i in range(3):
        book = tmp_path / f"cycle{i}.xlsx"
        run_cli(db, ["export", str(book)])
        books.append(cells(book))
        before = snapshot(db)
        r = run_cli(db, ["import", str(book)])
        assert r["review"] == []
        assert r["today_added"] == []  # membership never duplicates
        assert snapshot(db) == before  # the import changed nothing
    assert books[0] == books[1] == books[2]  # byte-stable cell content
    assert [t["name"] for t in run_cli(db, ["today"])["items"]] == [
        "Wire thermocouples", "Mount load cell",
    ]


def test_delete_project_with_completed_work_requires_force(page, board):
    """Completed work is the training data for estimation: deleting it must
    demand an explicit force, with the count spelled out."""
    lens = next(n for n in run_cli(board, ["ls"]) if n["name"] == "Lens spec")
    if lens["status"] != "done":
        run_cli(board, ["done", str(lens["id"])])
    row = tree_nodes(page, "project").filter(has_text="Harbour Lights")
    row.hover()
    row.locator('[data-testid="tree-delete"]').click()
    dlg = page.locator('[data-testid="confirm-dialog"]')
    expect(dlg).to_be_visible()
    expect(page.locator('[data-testid="confirm-completed-warning"]')).to_be_visible()
    expect(page.locator('[data-testid="confirm-completed-warning"]')).to_contain_text(
        "completed task"
    )
    ok = page.locator('[data-testid="confirm-ok"]')
    expect(ok).to_be_disabled()  # force is required, not a default
    page.locator('[data-testid="confirm-force"]').check()
    expect(ok).to_be_enabled()
    ok.click()
    expect(toast(page)).to_contain_text("Deleted project “Harbour Lights”")
    expect(tree(page)).not_to_contain_text("Harbour Lights")
    assert not any(
        n["name"] == "Harbour Lights" for n in run_cli(board, ["ls"])
    )
