"""Tests for the deterministic stage-1 Excel importer and template exporter.

Contract (from the template's README sheet):
- Columns are matched by header name; a file with only Project/Milestone/Goal/
  Task/Notes is fully valid (the sparse case).
- Kind from column position, order from row order, explicit Seq marks parallel
  ranks. No LLM anywhere in stage 1.
- 'Depends on' names prerequisite goals (or 'task: Name') from any sheet and is
  applied deterministically; 'Proposed: Depends on' kept in the file counts as
  accepted on import.
- Notes are never modified; notes that *smell* like dependency constraints are
  flagged for investigation (deterministic keyword heuristic).
- Ref ids are stable; matching by ref means renames never break links.
"""
from pathlib import Path

import pytest

from protracker.commands import CommandError, Commands
from protracker.importer import (
    _theme_palette,
    classify_health,
    classify_workbook,
    note_dependency_terms,
    read_workbook_rows,
)
from protracker.storage import Repository

FIXTURES = Path(__file__).parent / "fixtures"
SPARSE = FIXTURES / "tracker_sparse.xlsx"
NESTED = FIXTURES / "tracker_nested.xlsx"
TEMPLATE = FIXTURES / "tracker_template.xlsx"

PH = ("Project", "Milestone", "Goal", "Task", "Notes")
RICH = ("Project", "Milestone", "Goal", "Task", "Seq", "Depends on",
        "Deadline", "Est (min)", "Tags", "Notes", "Ref")

# The nested fixture is the enrichment answer key: 20 edges by (from, to) name.
ANSWER_KEY_EDGES = {
    ("Form business entity", "Secure permits & licenses"),
    ("Form business entity", "Accounting setup"),
    ("Form business entity", "Secure facility"),
    ("Secure facility", "Buildout"),
    ("Buildout", "Install roaster"),
    ("Install roaster", "First production roast"),
    ("Source green beans", "First production roast"),
    ("Secure permits & licenses", "First production roast"),
    ("Secure facility", "Cafe fit-out"),
    ("Lock production profiles", "Tasting with production coffee"),
    ("Lock production profiles", "Barista training on house coffee"),
    ("Cafe fit-out", "Soft open"),
    ("Menu development", "Soft open"),
    ("Hire & train staff", "Soft open"),
    ("Secure permits & licenses", "Soft open"),
    ("Business insurance", "Soft open"),
    ("Accounting setup", "Wholesale pricing"),
    ("First production roast", "Sales pipeline"),
    ("Brand identity", "Sales pipeline"),
    ("Wholesale pricing", "Sales pipeline"),
}


def R(*rows, width=None):
    """Rows as (rownum, padded tuple); width defaults to first row's length."""
    width = width or max(len(r) for r in rows)
    return [
        (i, tuple(list(r) + [None] * (width - len(r))))
        for i, r in enumerate(rows, start=1)
    ]


@pytest.fixture
def c():
    return Commands(Repository(":memory:"))


# --- pure classifier: sparse layout ---


def test_classify_minimal_sheet():
    plan = classify_workbook({
        "S": R(
            PH,
            ("P", None, None, None, "project note"),
            (None, "M"),
            (None, None, "G"),
            (None, None, None, "t1"),
            (None, None, None, "t2", "task note"),
        )
    })
    assert plan.review == []
    by_name = {n.name: n for n in plan.nodes}
    assert by_name["P"].kind == "project" and by_name["P"].path == ()
    assert by_name["M"].path == ("P",)
    assert by_name["G"].path == ("P", "M")
    t1, t2 = by_name["t1"], by_name["t2"]
    assert (t1.seq_index, t2.seq_index) == (1, 2)
    assert t1.seq_source == "assumed"  # row order is a guess
    assert t2.description == "task note"


def test_task_without_goal_after_project_goes_to_review():
    plan = classify_workbook({
        "S": R(PH, ("P",), (None, "M"), (None, None, None, "orphan task"))
    })
    assert [n.name for n in plan.nodes] == ["P", "M"]
    assert len(plan.review) == 1
    assert "goal" in plan.review[0]["reason"]


def test_task_with_no_project_at_all_is_planner():
    # covers the sparse Task/Notes planner sheet AND the rich Planner sheet
    plan = classify_workbook({
        "Planner": R(("Task", "Notes"), ("Pay taxes", "quarterly"))
    })
    assert plan.review == []
    pay = plan.nodes[0]
    assert pay.kind == "task" and pay.path == () and pay.seq_index is None
    assert pay.description == "quarterly"


def test_multiple_structural_cells_go_to_review():
    plan = classify_workbook({"S": R(PH, ("P",), (None, "M", "G both"))})
    assert [n.name for n in plan.nodes] == ["P"]
    assert "multiple" in plan.review[0]["reason"]


def test_note_only_row_goes_to_review():
    plan = classify_workbook({
        "S": R(PH, ("P",), (None, None, None, None, "floating note"))
    })
    assert len(plan.review) == 1


def test_non_tracker_sheet_is_skipped_not_reviewed():
    plan = classify_workbook({"README": R(("Once upon a time",), ("prose",))})
    assert plan.nodes == [] and plan.review == []
    assert plan.skipped[0]["sheet"] == "README"


# --- pure classifier: rich layout ---


def test_rich_columns_parse():
    plan = classify_workbook({
        "S": R(
            RICH,
            ("P", None, None, None, None, None, "2026-09-30", None, None, None, "PR"),
            (None, "M"),
            (None, None, "G", None, None, "Other goal", None, None, None, "needs stuff", "PR.m1.g"),
            (None, None, None, "t1", 1, None, None, 120, "desk;call", None, "PR.m1.g.t1"),
            (None, None, None, "t2", 2, "task: Elsewhere", "2026-08-01", "M", "deep", None, None),
        )
    })
    assert plan.review == []
    by_name = {n.name: n for n in plan.nodes}
    assert by_name["P"].deadline == "2026-09-30" and by_name["P"].ref == "PR"
    g = by_name["G"]
    assert g.depends_on == (("goal", "Other goal"),)
    assert g.ref == "PR.m1.g"
    t1 = by_name["t1"]
    assert t1.seq_index == 1 and t1.seq_source == "user"  # explicit Seq
    assert t1.est_minutes == 120 and t1.tags == ("desk", "call")
    t2 = by_name["t2"]
    assert t2.depends_on == (("task", "Elsewhere"),)
    assert t2.est_minutes == 45  # bucket M midpoint
    assert t2.deadline == "2026-08-01"


def test_equal_seq_cells_mean_parallel_rank():
    plan = classify_workbook({
        "S": R(
            RICH,
            ("P",), (None, "M"), (None, None, "G"),
            (None, None, None, "a", 1),
            (None, None, None, "b", 2),
            (None, None, None, "c", 2),
            (None, None, None, "d", 3),
        )
    })
    seqs = {n.name: n.seq_index for n in plan.nodes if n.kind == "task"}
    assert seqs == {"a": 1, "b": 2, "c": 2, "d": 3}


# --- deterministic note flagging ---


def test_dependency_smelling_notes_are_flagged():
    assert note_dependency_terms("needs the LLC") == ("need",)
    assert "same week" in note_dependency_terms(
        "fire and health scheduling can happen the same week"
    )
    assert note_dependency_terms("Rename this cell to your project name") == ()


def test_classifier_flags_notes_on_their_nodes():
    plan = classify_workbook({
        "S": R(
            PH,
            ("P",), (None, "M"),
            (None, None, "G1", None, "start after the entity stuff is done"),
            (None, None, None, "t1", "a lovely task"),
        )
    })
    assert len(plan.flags) == 1
    flag = plan.flags[0]
    assert flag["name"] == "G1" and "after" in flag["terms"]


def test_sparse_import_flags_all_dependency_notes(c):
    r = c.import_excel(str(SPARSE))
    flagged_names = {f["name"] for f in r["flagged"]}
    # every loose-dependency note in the sparse fixture must be surfaced
    assert {
        "Secure permits & licenses", "Accounting setup", "Secure facility",
        "Buildout", "Install roaster", "Source green beans",
        "First production roast", "Cafe fit-out",
        "Tasting with production coffee", "Barista training on house coffee",
        "Soft open", "Wholesale pricing", "Sales pipeline",
        "Schedule fire inspection", "Electrical & gas upgrades",
    } <= flagged_names
    # flags are suggestions only: no edges were invented
    assert c.list_dependencies() == []


# --- fixture ground truth ---


def test_sparse_fixture_counts(c):
    r = c.import_excel(str(SPARSE))
    assert r["review"] == [] and len(r["created"]) == 96
    assert len(c.list_nodes(kind="goal")) == 16
    assert len([
        t for t in c.list_nodes(kind="task") if t["parent_id"] is not None
    ]) == 63


def test_sparse_and_nested_hierarchies_identical():
    key = lambda p: {(n.kind, n.path, n.name) for n in p.nodes}
    sparse = classify_workbook(read_workbook_rows(str(SPARSE)))
    nested = classify_workbook(read_workbook_rows(str(NESTED)))
    assert key(sparse) == key(nested)


def test_nested_import_applies_answer_key_edges(c):
    r = c.import_excel(str(NESTED))
    assert r["review"] == []
    nodes = {n["id"]: n["name"] for n in c.list_nodes()}
    edges = {
        (nodes[d["from_id"]], nodes[d["to_id"]]) for d in c.list_dependencies()
    }
    assert edges == ANSWER_KEY_EDGES


def test_nested_import_metadata(c):
    c.import_excel(str(NESTED))
    task = next(
        n for n in c.list_nodes(kind="task")
        if n["name"] == "Choose business structure with accountant"
    )
    assert task["est_minutes"] == 120
    assert task["tags"] == ["desk", "call"]
    assert task["ref"] == "OPS.M1.entity.t1"
    assert task["seq_source"] == "user"  # explicit Seq column
    fire = next(
        n for n in c.list_nodes(kind="task")
        if n["name"] == "Schedule fire inspection"
    )
    health = next(
        n for n in c.list_nodes(kind="task")
        if n["name"] == "Schedule health inspection"
    )
    assert fire["seq_index"] == health["seq_index"] == 3  # parallel rank
    project = next(
        n for n in c.list_nodes(kind="project")
        if n["name"] == "Licensing & Compliance"
    )
    assert project["deadline"] == "2026-09-30"


def test_nested_reimport_is_idempotent(c):
    c.import_excel(str(NESTED))
    r = c.import_excel(str(NESTED))
    assert r["created"] == [] and r["updated"] == []
    assert r["dependencies_added"] == []
    assert r["unchanged"] == 96


def test_ref_match_survives_rename(c):
    c.import_excel(str(NESTED))
    llc = next(
        n for n in c.list_nodes(kind="task") if n["name"] == "File LLC formation"
    )
    c.update_node(llc["id"], name="File the LLC")
    r = c.import_excel(str(NESTED))  # ref OPS.M1.entity.t2 still matches
    assert r["created"] == []
    assert c.get_node(llc["id"])["node"]["name"] == "File LLC formation"


def test_unresolvable_depends_on_goes_to_review(c, tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    for _, row in R(
        RICH,
        ("P",), (None, "M"),
        (None, None, "G", None, None, "No Such Goal"),
        (None, None, None, "t1", 1),
    ):
        ws.append(list(row))
    path = str(tmp_path / "bad_dep.xlsx")
    wb.save(path)
    r = c.import_excel(path)
    assert any("No Such Goal" in str(e["values"]) for e in r["review"])
    assert c.list_dependencies() == []


def test_template_file_classifies_cleanly():
    plan = classify_workbook(read_workbook_rows(str(TEMPLATE)))
    assert {s["sheet"] for s in plan.skipped} == {"README"}
    names = {n.name for n in plan.nodes}
    assert "Example Project" in names
    # the example's prose proposal is parsed, resolution is deferred to apply
    ship = next(n for n in plan.nodes if n.name == "Ship the thing")
    assert ship.depends_on == (("goal", "Prepare the thing"),)
    assert ship.proposed == (("goal", "Permits goal (from Notes)"),)


# --- exporter round trip ---


def test_export_round_trip(c, tmp_path):
    c.import_excel(str(SPARSE))
    out = str(tmp_path / "enriched.xlsx")
    r = c.export_excel(out)
    assert r["nodes"] == 96

    plan = classify_workbook(read_workbook_rows(out))
    assert plan.review == []
    # hierarchy survives the round trip exactly
    key = lambda p: {(n.kind, n.path, n.name) for n in p.nodes}
    assert key(plan) == key(classify_workbook(read_workbook_rows(str(SPARSE))))

    # flagged notes surface in the Proposed column for human/LLM investigation
    rows = read_workbook_rows(out)
    all_cells = [
        str(v) for sheet in rows.values() for entry in sheet for v in entry[1] if v
    ]
    assert any("investigate" in v.lower() for v in all_cells)

    # re-import of our own export is clean and idempotent
    c2 = Commands(Repository(":memory:"))
    r2 = c2.import_excel(out)
    assert r2["review"] == []
    assert len(r2["created"]) == 96


def test_export_after_nested_import_writes_edges(c, tmp_path):
    c.import_excel(str(NESTED))
    out = str(tmp_path / "roundtrip.xlsx")
    c.export_excel(out)
    c2 = Commands(Repository(":memory:"))
    c2.import_excel(out)
    nodes = {n["id"]: n["name"] for n in c2.list_nodes()}
    edges = {
        (nodes[d["from_id"]], nodes[d["to_id"]]) for d in c2.list_dependencies()
    }
    assert edges == ANSWER_KEY_EDGES


def test_import_missing_file_is_structured_error(c):
    with pytest.raises(CommandError) as ei:
        c.import_excel("no_such_file.xlsx")
    assert ei.value.code == "not_found"


# --- multi-sheet workbooks: sheet identity, duplicate names, date typing ---
#
# Regression suite for the real-world workbook shape: one sheet per project,
# a preamble block copied unchanged from the template across every sheet,
# repeated sibling task names, and quarter/prose text in the date columns.

PRE = ("Project Milestone(s)", "Goal(s)", "Tasks", "Notes")


def _sheet(project_name, *body, header=PRE):
    """A preamble sheet: 'Project name' block, then header, then body rows."""
    return R(("Project name", project_name), header, *body, width=len(header))


def test_same_preamble_name_on_two_sheets_yields_two_projects():
    # The template ships a 'Project name' cell; copying the sheet and forgetting
    # to change it must not collapse two projects into one.
    plan = classify_workbook({
        "Alpha": _sheet("Shared Program", ("M1",), (None, "G1"), (None, None, "t1")),
        "Beta": _sheet("Shared Program", ("M2",), (None, "G2"), (None, None, "t2")),
    })
    projects = [n for n in plan.nodes if n.kind == "project"]
    assert [p.name for p in projects] == ["Alpha", "Beta"]
    # the collision is surfaced, never silent
    assert len(plan.review) == 2
    assert all("sheet title" in r["reason"] for r in plan.review)
    # children hang off their own sheet's project
    assert {n.name: n.path[0] for n in plan.nodes if n.kind == "milestone"} == {
        "M1": "Alpha", "M2": "Beta",
    }


def test_unique_preamble_name_still_beats_sheet_title():
    plan = classify_workbook({
        "Alpha": _sheet("Rocket Program", ("M1",), (None, "G1"), (None, None, "t1")),
        "Beta": _sheet("Coffee Shop", ("M2",), (None, "G2"), (None, None, "t2")),
    })
    projects = [n for n in plan.nodes if n.kind == "project"]
    assert [p.name for p in projects] == ["Rocket Program", "Coffee Shop"]
    assert plan.review == []


def test_repeated_sibling_task_names_are_distinct_nodes():
    plan = classify_workbook({
        "S": R(PH, ("P",), (None, "M"), (None, None, "G"),
               (None, None, None, "placeholder"),
               (None, None, None, "placeholder"),
               (None, None, None, "real task"),
               (None, None, None, "placeholder"))
    })
    occ = [(n.name, n.occurrence, n.seq_index)
           for n in plan.nodes if n.kind == "task"]
    assert occ == [
        ("placeholder", 1, 1), ("placeholder", 2, 2),
        ("real task", 1, 3), ("placeholder", 3, 4),
    ]


def test_repeated_task_names_import_as_three_nodes(c, tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    for col, head in enumerate(PH, start=1):
        ws.cell(row=1, column=col, value=head)
    ws["A2"] = "P"
    ws["B3"] = "M"
    ws["C4"] = "G"
    for i in range(3):
        ws.cell(row=5 + i, column=4, value="placeholder")
    out = str(tmp_path / "dup.xlsx")
    wb.save(out)

    r = c.import_excel(out)
    tasks = [n for n in c.list_nodes(kind="task")]
    assert len(tasks) == 3, "three rows must not collapse into one node"
    assert len(r["created"]) == 6  # project + milestone + goal + 3 tasks
    # and re-import is still idempotent: no duplication on a second pass
    r2 = c.import_excel(out)
    assert r2["created"] == []
    assert len(c.list_nodes(kind="task")) == 3


def test_duplicate_container_names_are_surfaced_for_review():
    plan = classify_workbook({
        "S": R(PH, ("P",), (None, "M"), (None, None, "G"),
               (None, None, None, "t1"), (None, None, "G"),
               (None, None, None, "t2"))
    })
    reasons = [r["reason"] for r in plan.review]
    assert any("duplicate goal" in r for r in reasons)


def test_non_date_text_does_not_land_in_a_date_column():
    plan = classify_workbook({
        "S": R(("Project", "Milestone", "Goal", "Task", "Start", "Finish"),
               ("P",), (None, "M"), (None, None, "G"),
               (None, None, None, "t1", "Q2", "5 weeks after start"))
    })
    t1 = next(n for n in plan.nodes if n.kind == "task")
    assert t1.deadline is None and t1.earliest_start is None
    # the value is preserved as an extra, not discarded
    assert ("Start", "Q2") in t1.extras
    assert ("Finish", "5 weeks after start") in t1.extras
    assert len(plan.review) == 2
    assert all("not a date" in r["reason"] for r in plan.review)


def test_iso_dates_still_populate_date_columns():
    plan = classify_workbook({
        "S": R(("Project", "Milestone", "Goal", "Task", "Start", "Finish"),
               ("P",), (None, "M"), (None, None, "G"),
               (None, None, None, "t1", "2026-08-01", "2026-09-01"))
    })
    t1 = next(n for n in plan.nodes if n.kind == "task")
    assert t1.earliest_start == "2026-08-01"
    assert t1.deadline == "2026-09-01"
    assert plan.review == []


# --- colour legend: quarter outlook (health), orthogonal to status ---
#
# Legend as written in the workbook:
#   Green  - on track for deadline
#   Yellow - tasks that have not begun but will be completed by the quarter
#   Red    - tasks that will not be done this quarter
#   Blue   - off track & Blue w/ Strikethrough - off track but was completed
#   Green & Strikethrough - completed (by deadline)
# Completion rides on strikethrough, so colour never writes status.

GREEN, YELLOW, RED, BLUE = "70AD47", "FFC000", "FF0000", "5B9BD5"


def RC(*rows, width=None):
    """Rows as (rownum, values, strikes, fills); each row is
    (values, strikes, fills) with strikes/fills defaulting to blank."""
    width = width or max(len(r[0]) for r in rows)

    def pad(seq, filler):
        return tuple(list(seq) + [filler] * (width - len(seq)))

    out = []
    for i, r in enumerate(rows, start=1):
        values, strikes, fills = (list(r) + [(), ()])[:3]
        out.append((i, pad(values, None), pad(strikes, False), pad(fills, None)))
    return out


def test_classify_health_covers_the_legend():
    assert classify_health(GREEN) == "on_track"
    assert classify_health(YELLOW) == "not_begun"
    assert classify_health(RED) == "wont_finish"
    assert classify_health(BLUE) == "off_track"


def test_classify_health_defaults_to_not_begun():
    # the user's rule: no colour means yellow
    for washed_out in (None, "", "FFFFFF", "000000", "A5A5A5"):
        assert classify_health(washed_out) == "not_begun"
    assert classify_health("nonsense") == "not_begun"


def test_classify_health_is_hue_based_not_exact_match():
    # a hand-picked green/red/blue still reads correctly
    assert classify_health("2E7D32") == "on_track"
    assert classify_health("C62828") == "wont_finish"
    assert classify_health("1565C0") == "off_track"


def test_fill_sets_health_and_strikethrough_still_sets_status():
    plan = classify_workbook({
        "S": RC(
            (PH,),
            (("P",),),
            ((None, "M"),),
            ((None, None, "G"),),
            ((None, None, None, "green done"), (False, False, False, True),
             (None, None, None, GREEN)),
            ((None, None, None, "blue done"), (False, False, False, True),
             (None, None, None, BLUE)),
            ((None, None, None, "red open"), (), (None, None, None, RED)),
            ((None, None, None, "uncoloured"),),
        )
    })
    tasks = {n.name: n for n in plan.nodes if n.kind == "task"}
    # green + strikethrough: completed by deadline
    assert (tasks["green done"].status, tasks["green done"].health) == (
        "done", "on_track")
    # blue + strikethrough: off track but was completed — both axes at once
    assert (tasks["blue done"].status, tasks["blue done"].health) == (
        "done", "off_track")
    assert (tasks["red open"].status, tasks["red open"].health) == (
        None, "wont_finish")
    assert (tasks["uncoloured"].status, tasks["uncoloured"].health) == (
        None, "not_begun")


def test_containers_carry_no_health():
    plan = classify_workbook({
        "S": RC((PH,), (("P",), (), (GREEN,)), ((None, "M"), (), (None, GREEN)),
                ((None, None, "G"), (), (None, None, GREEN)),
                ((None, None, None, "t1"),))
    })
    for n in plan.nodes:
        if n.kind != "task":
            assert n.health is None, f"{n.kind} must not carry a task colour"


def test_theme_palette_uses_the_index_order_not_the_xml_order(tmp_path):
    # <color theme="n"/> indexes lt1,dk1,lt2,dk2,accent1..6 — the first two are
    # swapped relative to how the tags appear in the theme XML. Getting this
    # wrong shifts every accent by one and silently mis-colours the workbook.
    import openpyxl

    out = str(tmp_path / "themed.xlsx")
    openpyxl.Workbook().save(out)
    palette = _theme_palette(openpyxl.load_workbook(out))
    assert len(palette) == 12
    assert palette[0] == "FFFFFF"  # lt1, not dk1
    assert palette[1] == "000000"  # dk1 — resolved from sysClr's lastClr


def test_no_theme_means_no_crash():
    class Themeless:
        loaded_theme = None

    assert _theme_palette(Themeless()) == []


def test_theme_indexed_fill_resolves_through_the_workbook_theme(tmp_path):
    # The real tracker paints with theme colours, not literal RGB, and each
    # workbook carries its own scheme: theme index 9 is green in the modern
    # Office theme but orange in the 2007 one. Resolution must therefore go
    # through *this* workbook's palette. Index 6 (accent3) is green here.
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.styles.colors import Color

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    for col, head in enumerate(PH, start=1):
        ws.cell(row=1, column=col, value=head)
    ws["A2"] = "P"
    ws["B3"] = "M"
    ws["C4"] = "G"
    cell = ws.cell(row=5, column=4, value="themed")
    green = Color(theme=6, tint=0.0)
    cell.fill = PatternFill("solid", start_color=green, end_color=green)
    out = str(tmp_path / "themed.xlsx")
    wb.save(out)

    palette = _theme_palette(openpyxl.load_workbook(out))
    assert classify_health(palette[6]) == "on_track", "fixture assumption"

    plan = classify_workbook(read_workbook_rows(out))
    themed = next(n for n in plan.nodes if n.name == "themed")
    assert themed.health == "on_track"


def test_health_and_completion_survive_an_export_round_trip(c, tmp_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    src = str(tmp_path / "src.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    for col, head in enumerate(PH, start=1):
        ws.cell(row=1, column=col, value=head)
    ws["A2"] = "P"
    ws["B3"] = "M"
    ws["C4"] = "G"
    for i, (name, colour, struck) in enumerate([
        ("on track", GREEN, False),
        ("off track done", BLUE, True),
        ("doomed", RED, False),
        ("plain", None, False),
    ]):
        cell = ws.cell(row=5 + i, column=4, value=name)
        if colour:
            cell.fill = PatternFill("solid", start_color=colour, end_color=colour)
        if struck:
            cell.font = Font(strikethrough=True)
    wb.save(src)

    c.import_excel(src)
    before = {n["name"]: (n["health"], n["status"])
              for n in c.list_nodes(kind="task")}
    assert before == {
        "on track": ("on_track", "active"),
        "off track done": ("off_track", "done"),
        "doomed": ("wont_finish", "active"),
        "plain": ("not_begun", "active"),
    }

    out = str(tmp_path / "out.xlsx")
    c.export_excel(out)
    c2 = Commands(Repository(":memory:"))
    c2.import_excel(out)
    after = {n["name"]: (n["health"], n["status"])
             for n in c2.list_nodes(kind="task")}
    assert after == before


# --- planner fields & Today membership round trip --------------------------
#
# remind, priority, followup_days, earliest_start and Today-list membership
# used to have no columns in the export, so export -> re-import silently
# dropped them. These pin the lossless round trip and its edges: idempotent
# re-import, tombstones outranking the file, and hand-authored values.


def clocked(day="2026-07-28"):
    """(commands, box): a Commands on an in-memory db with a settable clock."""
    box = [f"{day}T09:00:00"]
    return Commands(Repository(":memory:"), now=lambda: box[0]), box


def rich_board(c):
    """A board using every planner field: returns {name: id}."""
    p = c.add_node(kind="project", name="P")["created"]["id"]
    m = c.add_node(kind="milestone", name="M", parent_id=p)["created"]["id"]
    g = c.add_node(kind="goal", name="G", parent_id=m)["created"]["id"]
    a = c.add_node(kind="task", name="Calibrate rig", parent_id=g,
                   priority="high", followup_days=3)["created"]["id"]
    b = c.add_node(kind="task", name="Order stock", parent_id=g,
                   priority="low", earliest_start="2026-08-01")["created"]["id"]
    r = c.plan_followup(name="Renew cert", on_date="2026-08-04")
    rem = r["planned"]["id"]
    # Today: B first, then A — order must survive
    c.today_add(b)
    c.today_add(a)
    return {"a": a, "b": b, "rem": rem}


def test_planner_fields_survive_an_export_round_trip(tmp_path):
    c, _ = clocked()
    rich_board(c)
    out = str(tmp_path / "board.xlsx")
    c.export_excel(out)

    c2, _ = clocked()
    r = c2.import_excel(out)
    assert r["review"] == []
    by_name = {n["name"]: n for n in c2.list_nodes(kind="task")}
    assert by_name["Calibrate rig"]["priority"] == "high"
    assert by_name["Calibrate rig"]["followup_days"] == 3
    assert by_name["Order stock"]["priority"] == "low"
    assert by_name["Order stock"]["earliest_start"] == "2026-08-01"
    rem = by_name["Renew cert"]
    assert rem["remind"] == 1
    assert rem["earliest_start"] == "2026-08-04"
    # the reminder is waiting again, exactly as it was in the source db
    assert any(u["name"] == "Renew cert" for u in c2.upcoming())


def test_today_membership_and_order_survive_a_round_trip(tmp_path):
    c, _ = clocked()
    rich_board(c)
    out = str(tmp_path / "board.xlsx")
    c.export_excel(out)

    c2, _ = clocked()
    r = c2.import_excel(out)
    names = lambda cc: [t["name"] for t in cc.today()["items"]]
    assert names(c2) == names(c) == ["Order stock", "Calibrate rig"]
    assert len(r["today_added"]) == 2

    # same-file re-import adds nothing twice and changes nothing
    r2 = c2.import_excel(out)
    assert r2["today_added"] == []
    assert names(c2) == ["Order stock", "Calibrate rig"]


def test_today_tombstone_outranks_the_file(tmp_path):
    c, _ = clocked()
    ids = rich_board(c)
    out = str(tmp_path / "board.xlsx")
    c.export_excel(out)

    c2, _ = clocked()
    c2.import_excel(out)
    listed = {t["name"]: t["id"] for t in c2.today()["items"]}
    c2.today_remove(listed["Calibrate rig"])
    r = c2.import_excel(out)  # the file still says Today
    assert r["today_added"] == []
    assert [t["name"] for t in c2.today()["items"]] == ["Order stock"]


def test_done_tasks_never_land_on_today_from_a_file(tmp_path):
    c, _ = clocked()
    c.add_node(kind="project", name="P")
    out = str(tmp_path / "done.xlsx")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    header = ("Project", "Milestone", "Goal", "Task", "Today")
    for col, head in enumerate(header, start=1):
        ws.cell(row=1, column=col, value=head)
    ws["A2"] = "P2"
    ws["B3"] = "M"
    ws["C4"] = "G"
    from openpyxl.styles import Font

    done = ws.cell(row=5, column=4, value="finished long ago")
    done.font = Font(strikethrough=True)
    ws.cell(row=5, column=5, value=1)
    wb.save(out)

    r = c.import_excel(out)
    assert r["today_added"] == []
    assert c.today()["items"] == []


def test_hand_authored_planner_columns(tmp_path):
    """Case-insensitive priorities, 'yes'/'x' flags, and bad values kept as
    extras and surfaced — never written into a typed field."""
    import openpyxl

    out = str(tmp_path / "hand.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    header = ("Project", "Milestone", "Goal", "Task", "Priority",
              "Follow-up (days)", "Remind", "Today")
    for col, head in enumerate(header, start=1):
        ws.cell(row=1, column=col, value=head)
    ws["A2"] = "P"
    ws["B3"] = "M"
    ws["C4"] = "G"
    ws.append((None, None, None, "loud", "HIGH", 2, "yes", "x"))
    ws.append((None, None, None, "vague", "urgent", "soon", None, None))
    wb.save(out)

    c, _ = clocked()
    r = c.import_excel(out)
    by_name = {n["name"]: n for n in c.list_nodes(kind="task")}
    assert by_name["loud"]["priority"] == "high"
    assert by_name["loud"]["followup_days"] == 2
    assert by_name["loud"]["remind"] == 1
    assert [t["name"] for t in c.today()["items"]] == ["loud"]
    # invalid values: surfaced in review, node still created cleanly
    assert by_name["vague"]["priority"] is None
    assert by_name["vague"]["followup_days"] is None
    reasons = " | ".join(x["reason"] for x in r["review"])
    assert "'urgent'" in reasons and "'soon'" in reasons


def test_blank_planner_cells_never_clear_on_reimport(tmp_path):
    """An older-format export (no planner columns) merged over a rich board
    must not wipe the fields — absence is not an instruction."""
    c, _ = clocked()
    rich_board(c)
    out = str(tmp_path / "board.xlsx")
    c.export_excel(out)

    # strip the planner columns, keeping Ref so everything merges by identity
    import openpyxl

    wb = openpyxl.load_workbook(out)
    for ws in wb.worksheets:
        headers = [cell.value for cell in ws[1]]
        for name in ("Start", "Priority", "Follow-up (days)", "Remind", "Today"):
            idx = headers.index(name) + 1
            ws.delete_cols(idx)
            headers.pop(idx - 1)
    old = str(tmp_path / "old-format.xlsx")
    wb.save(old)

    c.import_excel(old, decisions={"P": "merge"})
    by_name = {n["name"]: n for n in c.list_nodes(kind="task")}
    assert by_name["Calibrate rig"]["priority"] == "high"
    assert by_name["Calibrate rig"]["followup_days"] == 3
    assert by_name["Renew cert"]["remind"] == 1
    assert [t["name"] for t in c.today()["items"]] == [
        "Order stock", "Calibrate rig",
    ]


# --- paths that arrive from a human ---------------------------------------
#
# The desktop app passes argv straight to the CLI so nothing typed into a
# dialog can be read as a shell command. That also means no shell strips the
# quotes Windows Explorer's "Copy as path" puts around a path, and they end up
# inside the filename: the extension check then fails on `.xlsx"`.


def test_quoted_paths_are_accepted(c):
    from protracker.commands import normalize_path

    assert normalize_path('"C:\a\b.xlsx"') == "C:\a\b.xlsx"
    assert normalize_path("'C:\a\b.xlsx'") == "C:\a\b.xlsx"
    assert normalize_path('  "C:\a\b.xlsx"  ') == "C:\a\b.xlsx"
    assert normalize_path("C:\a\b.xlsx") == "C:\a\b.xlsx"
    assert normalize_path("") == ""


def test_import_accepts_a_path_copied_with_quotes(c):
    r = c.import_excel(f'"{SPARSE}"')
    assert len(r["created"]) == 96


def test_import_of_a_non_workbook_is_a_structured_error(c, tmp_path):
    junk = tmp_path / "notes.txt"
    junk.write_text("this is not a workbook", encoding="utf8")
    with pytest.raises(CommandError) as ei:
        c.import_excel(str(junk))
    assert ei.value.code == "unreadable_workbook"
    assert "could not read" in ei.value.message


def test_import_of_a_truncated_workbook_is_a_structured_error(c, tmp_path):
    broken = tmp_path / "broken.xlsx"
    broken.write_bytes(SPARSE.read_bytes()[:400])  # valid name, corrupt content
    with pytest.raises(CommandError) as ei:
        c.import_excel(str(broken))
    assert ei.value.code == "unreadable_workbook"


def test_export_accepts_a_quoted_path(c, tmp_path):
    c.import_excel(str(SPARSE))
    out = tmp_path / "out.xlsx"
    r = c.export_excel(f'"{out}"')
    assert r["exported"] == str(out)
    assert out.exists()


def test_waits_and_repeats_survive_a_round_trip(tmp_path):
    """Spec 11.6: Wait reason and Repeat columns round-trip losslessly,
    and an unreadable Repeat value is kept as an extra, never coerced."""
    c, _ = clocked()
    ids = rich_board(c)
    c.wait(ids["b"], "2026-08-17", reason="pump lead time")
    c.plan_followup(name="Pay rent", on_date="2026-08-01",
                    repeat="monthly @date")
    out = str(tmp_path / "board.xlsx")
    c.export_excel(out)

    c2, _ = clocked()
    r = c2.import_excel(out)
    assert r["review"] == []
    by_name = {n["name"]: n for n in c2.list_nodes(kind="task")}
    assert by_name["Order stock"]["wait_reason"] == "pump lead time"
    assert by_name["Order stock"]["earliest_start"] == "2026-08-17"
    rent = by_name["Pay rent"]
    assert rent["repeat"] is not None
    from protracker import recurrence

    assert recurrence.format_rule(recurrence.loads(rent["repeat"])) == \
        "monthly @date"
    # the wait renders as an external blocker on the fresh database too
    b = c2.get_node(by_name["Order stock"]["id"])["blockers"][0]
    assert b["type"] == "external" and b["reason"] == "pump lead time"


def test_hand_authored_repeat_column(tmp_path):
    import openpyxl

    out = str(tmp_path / "rep.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    header = ("Project", "Milestone", "Goal", "Task", "Repeat", "Wait reason")
    for col, head in enumerate(header, start=1):
        ws.cell(row=1, column=col, value=head)
    ws["A2"] = "P"
    ws["B3"] = "M"
    ws["C4"] = "G"
    ws.append((None, None, None, "water plants", "every 3d", None))
    ws.append((None, None, None, "junk rule", "sometimes", None))
    ws.append((None, None, None, "await pump", None, "vendor"))
    wb.save(out)

    c, _ = clocked()
    r = c.import_excel(out)
    by_name = {n["name"]: n for n in c.list_nodes(kind="task")}
    assert by_name["water plants"]["repeat"] is not None
    assert by_name["junk rule"]["repeat"] is None
    assert by_name["await pump"]["wait_reason"] == "vendor"
    reasons = " | ".join(x["reason"] for x in r["review"])
    assert "'sometimes'" in reasons and "not a repeat rule" in reasons


def test_steps_and_links_survive_a_round_trip(tmp_path):
    c, _ = clocked()
    ids = rich_board(c)
    c.step_add(ids["a"], "cut; then measure")  # semicolon in a step name
    s2 = c.step_add(ids["a"], "deburr")["steps"][1]
    c.step_tick(s2["id"])
    c.link_add(ids["a"], "C:/data/run7.csv", label="run 7 data")
    c.link_add(ids["a"], "https://vendor.example/quote")
    out = str(tmp_path / "board.xlsx")
    c.export_excel(out)

    c2, _ = clocked()
    r = c2.import_excel(out)
    assert r["steps_added"] == 2 and r["links_added"] == 2
    by_name = {n["name"]: n for n in c2.list_nodes(kind="task")}
    a2 = by_name["Calibrate rig"]["id"]
    steps = c2.get_node(a2)["steps"]
    assert [(s["name"], s["done"]) for s in steps] == [
        ("cut; then measure", 0), ("deburr", 1),
    ]
    node = c2.get_node(a2)["node"]
    assert node["links"] == [
        {"label": "run 7 data", "href": "C:/data/run7.csv"},
        {"label": "https://vendor.example/quote",
         "href": "https://vendor.example/quote"},
    ]
    # idempotent: a second import adds nothing
    r2 = c2.import_excel(out)
    assert r2["steps_added"] == 0 and r2["links_added"] == 0
    # a tick in the file is a statement: tick locally, re-import never
    # unticks (the export says [x] now anyway); untick locally, the file's
    # [x] re-asserts it
    first = c2.get_node(a2)["steps"][1]
    c2.step_tick(first["id"], done=False)
    c2.import_excel(out)
    assert c2.get_node(a2)["steps"][1]["done"] == 1
