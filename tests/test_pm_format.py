"""The project manager's workbook format — the deliverable.

Its shape is fixed by an external obligation: the user must hand this file to
their project manager and cannot change its layout. So these tests pin the
layout itself (preamble, colour legend, header row, the eleven column names in
order) and the two round trips that matter:

- a file in this format imports and re-exports structurally unchanged;
- a tracker authored entirely through verbs — no import — exports complete,
  including the colour axis and completion dates.

The fixture is synthetic. The real tracker is unpublished research data and
stays out of the repository (`.gitignore` excludes workbooks by default).
"""
from __future__ import annotations

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill

from protracker.commands import CommandError, Commands
from protracker.exporter import PM_HEADER, PM_HEADER_ROW, PM_LEGEND
from protracker.storage import Repository

GREEN, YELLOW, RED, BLUE = "70AD47", "FFC000", "FF0000", "5B9BD5"


def node(c, nid):
    """`get_node` wraps the record; these tests want the fields."""
    return c.get_node(nid)["node"]


def state(c, nid):
    return c.get_node(nid)["state"]


def new(c, kind, name, **kw):
    """add_node takes only the classic fields; anything else is a follow-up
    update, which is also how the editor will do it."""
    extra = {k: kw.pop(k) for k in list(kw) if k not in (
        "parent_id", "description", "deadline", "earliest_start", "seq_index",
        "weight", "est_minutes", "tags", "priority", "followup_days")}
    nid = c.add_node(kind, name, **kw)["created"]["id"]
    if extra:
        c.update_node(nid, **extra)
    return nid


@pytest.fixture
def c():
    return Commands(Repository(":memory:"))


def make_pm_workbook(path):
    """The required layout, with invented tissue-engineering content: preamble
    rows 1-4, legend in column F, header on row 7, eleven columns, colours and
    strikethrough carrying health and completion."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fibrous Composites"
    ws["A1"] = "Project name"; ws["B1"] = "Fibrous Composites"
    ws["A2"] = "Tier Level"; ws["B2"] = "1"
    ws["A3"] = "Project start date"; ws["B3"] = "N/A"
    ws["A4"] = "Project finish date"; ws["B4"] = "2026-12-31"
    for row, text in PM_LEGEND:
        ws.cell(row=row, column=6, value=text)
    for col, head in enumerate(PM_HEADER, start=1):
        ws.cell(row=PM_HEADER_ROW, column=col, value=head)

    rows = [
        # (col, name, colour, struck, extras{})
        (1, "Identify an ideal hydrogel", None, False, {}),
        (2, "Chitogel + EDC/NHS characterisation", GREEN, True, {}),
        (3, "Fabricate chitogel cylinders", GREEN, True,
         {6: "2026-07-01", 7: "2026-07-10"}),
        (3, "Compressive mechanical testing", YELLOW, False,
         {5: "least brittle formulation identified",
          9: "construct was very swollen", 10: "repeat with lower EDC"}),
        (3, "Histology", BLUE, False, {8: "high"}),
        (2, "Pure collagen hydrogel", RED, False, {}),
        (3, "Cast 1% / 0.6% / 0.3% w/v gels", None, False, {}),
    ]
    r = PM_HEADER_ROW
    for col, name, colour, struck, extras in rows:
        r += 1
        cell = ws.cell(row=r, column=col, value=name)
        if colour:
            cell.fill = PatternFill("solid", start_color=colour, end_color=colour)
        if struck:
            cell.font = Font(strikethrough=True)
        for idx, value in extras.items():
            ws.cell(row=r, column=idx + 1, value=value)

    stats = wb.create_sheet("Stats sheet example")
    stats["A1"] = "* Stats are use for productivity *"
    wb.save(path)


def profile(path):
    """(sheets, header row cells, struck cells, filled cells) — the structural
    fingerprint an export has to reproduce. Coordinates are compared, not just
    counts, so a colour landing on the wrong row is caught."""
    wb = openpyxl.load_workbook(path)
    out = {}
    for ws in wb.worksheets:
        header = [c.value for c in ws[PM_HEADER_ROW]]
        struck, filled = set(), {}
        for row in ws.iter_rows(min_row=PM_HEADER_ROW + 1):
            for cell in row:
                if cell.value is not None and cell.font and cell.font.strike:
                    struck.add(cell.coordinate)
                if cell.fill and cell.fill.fill_type == "solid":
                    filled[cell.coordinate] = cell.fill.fgColor.rgb
        out[ws.title] = (header, struck, filled)
    return out


# --- the layout itself -----------------------------------------------------


def test_export_writes_the_required_layout(c, tmp_path):
    new(c, "project", "Fibrous Composites", tier_level="2",
        deadline="2026-12-31")
    out = str(tmp_path / "pm.xlsx")
    c.export_pm(out)

    ws = openpyxl.load_workbook(out)["Fibrous Composites"]
    assert [ws.cell(row=i, column=1).value for i in range(1, 5)] == [
        "Project name", "Tier Level", "Project start date",
        "Project finish date",
    ]
    assert ws["B1"].value == "Fibrous Composites"
    assert ws["B2"].value == "2"
    assert ws["B3"].value == "N/A"          # no start date set
    assert ws["B4"].value == "2026-12-31"
    for row, text in PM_LEGEND:
        assert ws.cell(row=row, column=6).value == text
    assert [c_.value for c_ in ws[PM_HEADER_ROW]][:len(PM_HEADER)] == list(PM_HEADER)
    assert len(PM_HEADER) == 11


def test_export_reports_what_the_format_cannot_carry(c, tmp_path):
    """The format has no Ref/Seq/Depends on/Est column, so the write is lossy.
    It must say so — silence is how 31 cells went missing on import."""
    new(c, "task", "pay invoice")  # a planner task has no place at all
    res = c.export_pm(str(tmp_path / "pm.xlsx"))
    assert res["format"] == "pm"
    joined = " ".join(res["omitted"])
    assert "planner task" in joined
    assert "dependencies" in joined


# --- round trip: a file in this format -------------------------------------


def test_pm_file_imports_and_re_exports_unchanged(c, tmp_path):
    src = str(tmp_path / "src.xlsx")
    make_pm_workbook(src)
    c.import_excel(src)
    out = str(tmp_path / "out.xlsx")
    c.export_pm(out)

    before, after = profile(src), profile(out)
    # the stats sheet is not a tracker sheet and is not reproduced
    assert set(after) == {"Fibrous Composites"}
    b, a = before["Fibrous Composites"], after["Fibrous Composites"]
    assert a[0] == b[0], "header row must match exactly"
    assert a[1] == b[1], "struck cells must land on the same coordinates"
    assert a[2] == b[2], "fills must land on the same coordinates and colours"


def test_the_four_previously_dropped_columns_survive(c, tmp_path):
    """Success Criteria and Trouble shooting Comments were absent from
    COLUMN_ROLES, so 31 cells of the real tracker were discarded on import."""
    src = str(tmp_path / "src.xlsx")
    make_pm_workbook(src)
    c.import_excel(src)
    by_name = {n["name"]: n for n in c.list_nodes()}
    task = by_name["Compressive mechanical testing"]
    assert task["success_criteria"] == "least brittle formulation identified"
    assert task["troubleshooting"] == "construct was very swollen"
    assert task["description"] == "repeat with lower EDC"
    assert by_name["Fibrous Composites"]["tier_level"] == "1"

    out = str(tmp_path / "out.xlsx")
    c.export_pm(out)
    c2 = Commands(Repository(":memory:"))
    c2.import_excel(out)
    round_tripped = {n["name"]: n for n in c2.list_nodes()}
    again = round_tripped["Compressive mechanical testing"]
    assert again["success_criteria"] == task["success_criteria"]
    assert again["troubleshooting"] == task["troubleshooting"]


def test_yellow_is_distinguishable_from_uncoloured(c, tmp_path):
    """Both read as "not begun", but only one is a statement. Conflating them
    made an authored yellow impossible to export."""
    src = str(tmp_path / "src.xlsx")
    make_pm_workbook(src)
    c.import_excel(src)
    by_name = {n["name"]: n for n in c.list_nodes()}
    assert by_name["Compressive mechanical testing"]["health"] == "not_begun"
    assert by_name["Cast 1% / 0.6% / 0.3% w/v gels"]["health"] is None

    out = str(tmp_path / "out.xlsx")
    c.export_pm(out)
    fills = profile(out)["Fibrous Composites"][2]
    # openpyxl reports fills as 8-char ARGB ('00FFC000')
    assert any(v.endswith(YELLOW) for v in fills.values())
    ws = openpyxl.load_workbook(out)["Fibrous Composites"]
    plain = [
        cell for row in ws.iter_rows(min_row=PM_HEADER_ROW + 1) for cell in row
        if cell.value == "Cast 1% / 0.6% / 0.3% w/v gels"
    ][0]
    assert plain.fill.fill_type != "solid", "an uncoloured row stays unpainted"


# --- round trip: a tracker authored in the app, never imported -------------


def test_authoring_from_empty_exports_complete(c, tmp_path):
    """The user's actual workflow: build the tracker in the app, then hand the
    workbook in. Nothing covered this before, and `health` was unwritable by
    any verb, so an authored tracker exported with no colours at all."""
    p = new(c, "project", "Looped Ligament", tier_level="1",
            deadline="2026-11-30")
    m = new(c, "milestone", "Culture around static posts", parent_id=p)
    g = new(c, "goal", "Assess scaffold contraction", parent_id=m)
    t1 = new(c, "task", "Seed ESMSCs on genipin-CX scaffolds", parent_id=g)
    t2 = new(c, "task", "Confocal live/dead", parent_id=g)

    c.update_node(t1, health="on_track", success_criteria="scaffolds contract",
                  team_lead="K", responsible_party="K")
    c.complete_task(t1)
    c.update_node(t1, completed_at="2026-07-14")
    c.update_node(t2, health="not_begun", troubleshooting="needs confocal slot")

    out = str(tmp_path / "pm.xlsx")
    c.export_pm(out)

    c2 = Commands(Repository(":memory:"))
    c2.import_excel(out)
    by_name = {n["name"]: n for n in c2.list_nodes()}
    done = by_name["Seed ESMSCs on genipin-CX scaffolds"]
    assert done["status"] == "done"           # strikethrough
    assert done["health"] == "on_track"       # green
    assert done["success_criteria"] == "scaffolds contract"
    assert done["team_lead"] == "K"
    open_task = by_name["Confocal live/dead"]
    assert open_task["health"] == "not_begun"
    assert open_task["troubleshooting"] == "needs confocal slot"
    assert by_name["Looped Ligament"]["tier_level"] == "1"


def test_completion_date_is_editable_but_never_clearable_while_done(c):
    """The user wants to see and adjust the date a row went
    green-with-strikethrough. Clearing it on a finished node would export as
    struck-but-undated, so that is refused rather than written."""
    g = _goal(c)
    t = new(c, "task", "Histology", parent_id=g)
    c.complete_task(t)
    c.update_node(t, completed_at="2026-06-01")
    assert node(c, t)["completed_at"] == "2026-06-01"
    with pytest.raises(CommandError) as exc:
        c.update_node(t, completed_at=None)
    assert exc.value.code == "invalid_field"


def test_health_is_validated(c):
    g = _goal(c)
    t = new(c, "task", "SEM imaging", parent_id=g)
    with pytest.raises(CommandError) as exc:
        c.update_node(t, health="chartreuse")
    assert exc.value.code == "invalid_field"
    c.update_node(t, health="off_track")
    c.update_node(t, health=None)
    assert node(c, t)["health"] is None


# --- pivots: finishing a goal without exhausting it ------------------------


def _goal(c):
    p = new(c, "project", "P")
    m = new(c, "milestone", "M", parent_id=p)
    return new(c, "goal", "G", parent_id=m)


def test_finishing_a_goal_takes_its_unrun_tasks_out_of_ready(c):
    """"I got what I needed from it, mark the goal done." Before this, a
    container had no done state and the nine tasks under it had to be dropped
    one at a time."""
    g = _goal(c)
    ids = [new(c, "task", f"t{i}", parent_id=g) for i in range(3)]
    assert [r["id"] for r in c.ready()] == [ids[0]]

    c.update_node(g, status="done")
    assert c.ready() == []
    # 'moot', not 'dropped': the work did not fail, it stopped being needed
    assert state(c, ids[0]) == "moot"
    assert node(c, ids[0])["status"] == "active"
    # and the goal is complete even though nothing under it was finished
    assert node(c, g)["completed_at"] is not None

    c.update_node(g, status="active")
    assert [r["id"] for r in c.ready()] == [ids[0]], "reopening restores them"
    assert node(c, g)["completed_at"] is None


def test_abandoned_goal_exports_red_and_done_goal_exports_struck(c, tmp_path):
    g_done = _goal(c)
    m = node(c, g_done)["parent_id"]
    g_drop = new(c, "goal", "pivoted away", parent_id=m)
    c.update_node(g_done, status="done", health="on_track")
    c.update_node(g_drop, status="abandoned")

    out = str(tmp_path / "pm.xlsx")
    c.export_pm(out)
    ws = openpyxl.load_workbook(out)["P"]
    cells = {
        cell.value: cell
        for row in ws.iter_rows(min_row=PM_HEADER_ROW + 1) for cell in row
        if cell.value
    }
    assert cells["G"].font.strike is True
    assert cells["G"].fill.fgColor.rgb.endswith(GREEN)
    # a pivot is not a completion: red ("will not be done"), not struck
    assert cells["pivoted away"].fill.fgColor.rgb.endswith(RED)
    assert not cells["pivoted away"].font.strike
