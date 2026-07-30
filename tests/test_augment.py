"""Tests for the LLM augmentation pipeline and the offline graph generator.

The LLM is stubbed with an injected transport: the pipeline itself must be
fully deterministic and offline. Also covers the legacy real-tracker format
(preamble block, row-6 header, sheet-as-project, strikethrough completion,
extra context columns).
"""
import json
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font

from protracker import augment
from protracker.commands import Commands
from protracker.importer import classify_workbook, read_workbook_rows
from protracker.storage import Repository

FIXTURES = Path(__file__).parent / "fixtures"
SPARSE = FIXTURES / "tracker_sparse.xlsx"


# ---------------------------------------------------------------- real format


def make_real_format(path: Path, extra_column: bool = False):
    """Mimics 'real Project Tracker Template.xlsx': preamble rows, header on
    row 6, no Project column, strikethrough = done.

    With `extra_column`, appends a column the importer does not recognise, so
    the extras-and-review path stays exercised now that every column of the
    real template is mapped."""
    tail = ["Vendor"] if extra_column else []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rocket"
    ws["A1"] = "Project name"; ws["B1"] = "Rocket Program"
    ws["A2"] = "Tier Level"; ws["B2"] = 1
    ws["A3"] = "Project start date"
    ws["A4"] = "Project finish date"; ws["B4"] = "2026-12-31"
    ws["F1"] = "Green - on track for deadline"
    ws.append([])  # row 5 spacer
    ws.append(["Project Milestone(s)", "Goal(s)", "Tasks", "Team Lead",
               "Responsible Party", "Success Criteria", "Start", "Finish",
               "Priority", "Trouble shooting Comments", "Notes"] + tail)
    ws.append(["Design phase", None, None, None, None, None, None,
               "2026-09-01", None, None, None])
    ws.append([None, "Engine spec", None, "Alice", None, None, None, None,
               None, None, "needs the fuel study first"])
    ws.append([None, None, "Draft requirements", "Alice", "Bob", "signed off",
               "2026-08-01", "2026-08-15", "High",
               "blocked until vendor call", None]
              + (["Acme Ltd"] if extra_column else []))
    ws.append([None, None, "Review with team", None, None, None, None, None,
               None, None, None])
    done_row = ws.max_row
    ws.cell(row=done_row, column=3).font = Font(strikethrough=True)

    stats = wb.create_sheet("Stats sheet example")
    stats["A1"] = "* Stats are use for productivity *"
    stats["A4"] = "Projects"; stats["B4"] = "Completed Tasks"
    wb.save(path)


def test_real_format_classification(tmp_path):
    path = tmp_path / "real.xlsx"
    make_real_format(path)
    plan = classify_workbook(read_workbook_rows(str(path)))
    assert plan.review == []
    assert {s["sheet"] for s in plan.skipped} == {"Stats sheet example"}
    by_name = {n.name: n for n in plan.nodes}
    project = by_name["Rocket Program"]  # from the preamble, not sheet title
    assert project.kind == "project"
    assert project.deadline == "2026-12-31"
    assert by_name["Design phase"].kind == "milestone"
    assert by_name["Design phase"].deadline == "2026-09-01"  # Finish column
    goal = by_name["Engine spec"]
    assert goal.kind == "goal"
    assert goal.path == ("Rocket Program", "Design phase")
    assert goal.description == "needs the fuel study first"
    draft = by_name["Draft requirements"]
    assert draft.kind == "task"
    assert draft.deadline == "2026-08-15"
    assert draft.earliest_start == "2026-08-01"
    # every column of the real template is a first-class field, not an extra:
    # these four used to be silently discarded
    assert draft.extras == ()
    assert draft.team_lead == "Alice"
    assert draft.responsible_party == "Bob"
    assert draft.success_criteria == "signed off"
    assert draft.troubleshooting == "blocked until vendor call"
    assert by_name["Engine spec"].team_lead == "Alice"  # containers too
    assert project.tier_level == "1"  # preamble
    review = by_name["Review with team"]
    assert review.status == "done"  # strikethrough
    assert draft.status is None


def test_real_format_imports_and_flags(tmp_path):
    path = tmp_path / "real.xlsx"
    make_real_format(path)
    c = Commands(Repository(":memory:"))
    r = c.import_excel(str(path))
    assert r["review"] == []
    done = next(
        n for n in c.list_nodes(kind="task") if n["name"] == "Review with team"
    )
    assert done["status"] == "done"
    flagged = {f["name"] for f in r["flagged"]}
    assert "Engine spec" in flagged  # "needs the fuel study first"


# ---------------------------------------------------------------- pipeline


def stub_transport(system, user):
    """Deterministic stand-in for an LLM."""
    assert "Output contract" in system  # instructions doc made it into the call
    assert "Source green beans" in user  # tracker outline made it in
    return {
        "dependencies": [
            {"from": "Form business entity", "from_kind": "goal",
             "to": "Secure permits & licenses", "to_kind": "goal",
             "reason": "note: start after the entity stuff is done"},
            {"from": "Secure permits & licenses", "from_kind": "goal",
             "to": "Form business entity", "to_kind": "goal",
             "reason": "bogus reverse edge"},
            {"from": "No Such Goal", "from_kind": "goal",
             "to": "Soft open", "to_kind": "goal",
             "reason": "hallucinated"},
        ],
        "parallel_groups": [
            {"tasks": ["Schedule fire inspection", "Schedule health inspection"],
             "reason": "note: same week"},
        ],
        "estimates": [
            {"task": "Obtain EIN", "bucket": "S", "reason": "simple form"},
            {"task": "Pass inspections & receive permits", "bucket": "SPLIT",
             "reason": "multi-visit"},
        ],
    }


@pytest.fixture(scope="module")
def result(tmp_path_factory):
    out = tmp_path_factory.mktemp("augment_out")
    return augment.run(str(SPARSE), str(out), transport=stub_transport), out


def test_pipeline_applies_valid_and_rejects_invalid(result):
    r, _ = result
    applied = r["proposals"]["applied"]
    rejected = r["proposals"]["rejected"]
    assert [d["to"] for d in applied["dependencies"]] == ["Secure permits & licenses"]
    reasons = " ".join(x["reason"] for x in rejected)
    assert "cycle" in reasons and "unresolved" in reasons
    assert len(rejected) == 2


def test_pipeline_parallel_and_estimates(result):
    r, out = result
    c = Commands(Repository(str(Path(r["outputs"]["db"]))))
    tasks = {t["name"]: t for t in c.list_nodes(kind="task")}
    assert (
        tasks["Schedule fire inspection"]["seq_index"]
        == tasks["Schedule health inspection"]["seq_index"]
    )
    ein = tasks["Obtain EIN"]
    assert ein["est_minutes"] == 15 and ein["est_source"] == "llm"
    # SPLIT recorded but no minutes invented
    assert tasks["Pass inspections & receive permits"]["est_minutes"] is None
    split = [e for e in r["proposals"]["applied"]["estimates"] if e["bucket"] == "SPLIT"]
    assert split


def test_pipeline_outputs_exist_and_are_offline(result):
    r, _ = result
    for key in ("xlsx", "html", "report"):
        assert Path(r["outputs"][key]).exists()
    html = Path(r["outputs"]["html"]).read_text(encoding="utf-8")
    assert "__DATA__" not in html and "__SOURCE__" not in html
    assert "Form business entity" in html
    assert "claude.ai" not in html  # fully local artifact
    report = Path(r["outputs"]["report"]).read_text(encoding="utf-8")
    assert "Rejected proposals" in report


def test_llm_deps_land_in_proposed_column(result):
    r, _ = result
    wb = openpyxl.load_workbook(r["outputs"]["xlsx"])
    ws = wb["Licensing & Compliance"]
    header = [c.value for c in ws[1]]
    dep_col = header.index("Depends on")
    prop_col = header.index("Proposed: Depends on")
    row = next(
        row for row in ws.iter_rows(min_row=2)
        if row[2].value == "Secure permits & licenses"
    )
    assert row[prop_col].value == "Form business entity"  # pending review
    assert row[dep_col].value is None  # not canonicalized


def test_dry_run_never_calls_llm(tmp_path):
    def exploding_transport(system, user):  # pragma: no cover
        raise AssertionError("transport must not be called in dry run")

    r = augment.run(
        str(SPARSE), str(tmp_path / "dry"), transport=exploding_transport,
        dry_run=True,
    )
    assert r["proposals"]["applied"]["dependencies"] == []
    assert Path(r["outputs"]["html"]).exists()


def test_emit_prompt_writes_query_and_calls_no_llm(tmp_path):
    def exploding_transport(system, user):  # pragma: no cover
        raise AssertionError("transport must not be called with --emit-prompt")

    r = augment.run(
        str(SPARSE), str(tmp_path / "p"), transport=exploding_transport,
        emit_prompt=True,
    )
    system = Path(r["outputs"]["prompt_system"]).read_text(encoding="utf-8")
    user = Path(r["outputs"]["prompt_user"]).read_text(encoding="utf-8")
    assert "Output contract" in system
    assert "Source green beans" in user
    assert "proposals.json" in r["next"]


def test_outline_includes_extras(tmp_path):
    path = tmp_path / "real.xlsx"
    make_real_format(path, extra_column=True)
    r = augment.run(str(path), str(tmp_path / "out"), dry_run=True)
    c = Commands(Repository(r["outputs"]["db"]))
    plan = classify_workbook(read_workbook_rows(str(path)))
    by_key = {(s.kind, s.name): s.extras for s in plan.nodes if s.extras}
    extras_by_ref = {
        n["ref"]: by_key[(n["kind"], n["name"])]
        for n in c.list_nodes() if (n["kind"], n["name"]) in by_key
    }
    outline = augment.build_outline(c, [], extras_by_ref)
    assert 'Vendor: "Acme Ltd"' in outline


def test_unrecognised_column_is_reported_not_dropped(tmp_path):
    """An unmapped column with content must be said out loud. Silence here is
    how 31 cells of the user's real tracker went missing."""
    path = tmp_path / "real.xlsx"
    make_real_format(path, extra_column=True)
    plan = classify_workbook(read_workbook_rows(str(path)))
    reasons = [r["reason"] for r in plan.review]
    assert any("'Vendor'" in r and "not imported" in r for r in reasons), reasons
